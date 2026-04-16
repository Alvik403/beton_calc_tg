 rom __future__ import annotations

from pathlib import Path
from typing import Iterable

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.web as web_module


@pytest.fixture(autouse=True)
def isolate_profile_state(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("TESTING", "1")
    monkeypatch.setattr(web_module, "PROFILES_PATH", tmp_path / "web_profiles.json")
    monkeypatch.setattr(web_module, "JBI_PROFILES_PATH", tmp_path / "web_profiles_jbi.json")
    web_module._last_request_per_ip.clear()
    yield
    web_module._last_request_per_ip.clear()


@pytest.fixture
def client():
    with TestClient(web_module.app) as test_client:
        yield test_client


@pytest.fixture
def config_password_header():
    return {"X-Config-Password": web_module.CONFIG_PASSWORD}


@pytest.fixture
def make_workbook(tmp_path: Path):
    def _make_workbook(
        rows: Iterable[dict[str, object]],
        *,
        filename: str = "balances.xlsx",
        saldo_col: int = 3,
        include_saldo_header: bool = True,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        if include_saldo_header:
            ws.cell(row=1, column=saldo_col, value="Сальдо на конец периода")
        current_row = 2
        for item in rows:
            col = int(item.get("column", 1))
            ws.cell(row=current_row, column=col, value=item["name"])
            ws.cell(row=current_row + 1, column=saldo_col, value=item.get("quantity", 0))
            current_row += 2
        path = tmp_path / filename
        wb.save(path)
        return path

    return _make_workbook
