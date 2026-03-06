from __future__ import annotations

import asyncio
import tempfile
from pathlib import Path
from decimal import Decimal

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart
import re

from aiogram.types import FSInputFile, Message

from app.calculator import Recipe, calculate_max_cubic_meters
from app.config import (
    get_bot_token,
    load_materials_config,
    load_prices_config,
    load_recipes_config,
)
from app.excel_parser import MaterialConfig, extract_balances


def _money(volume_m3: Decimal, price_per_m3: float) -> Decimal:
    return Decimal(str(volume_m3)) * Decimal(str(price_per_m3))


def _load_materials() -> list[MaterialConfig]:
    raw = load_materials_config()
    materials: list[MaterialConfig] = []
    for item in raw:
        name = item.get("name", "").strip()
        aliases = [a for a in item.get("aliases", []) if a]
        if name and aliases:
            materials.append(MaterialConfig(name=name, aliases=aliases))
    return materials


def _load_recipes() -> list[Recipe]:
    raw = load_recipes_config()
    recipes: list[Recipe] = []
    for item in raw:
        name = item.get("name", "").strip()
        materials = item.get("materials", {})
        if name and materials:
            recipes.append(Recipe(name=name, materials=materials))
    return recipes


def _normalize_name(text: str) -> str:
    text = text.strip().lower().replace("ё", "е")
    text = text.replace("в", "b").replace("з", "3")
    text = re.sub(r"\s+", " ", text)
    return text


def _load_prices() -> dict[str, dict[str, float]]:
    raw = load_prices_config()
    prices: dict[str, dict[str, float]] = {}
    for item in raw:
        name = item.get("name", "").strip()
        if not name:
            continue
        prices[_normalize_name(name)] = {
            "no_delivery_no_vat": float(item.get("no_delivery_no_vat", 0) or 0),
            "no_delivery_vat_22": float(item.get("no_delivery_vat_22", 0) or 0),
            "pickup_no_vat": float(item.get("pickup_no_vat", 0) or 0),
            "pickup_vat_22": float(item.get("pickup_vat_22", 0) or 0),
        }
    return prices


def _build_output_dataframe(
    recipes: list[Recipe], balances: dict[str, float]
) -> pd.DataFrame:
    all_materials = []
    for recipe in recipes:
        for material in recipe.materials:
            if material not in all_materials:
                all_materials.append(material)

    rows = []
    for recipe in recipes:
        max_m3, required = calculate_max_cubic_meters(recipe, balances)
        row = {"Наименование": recipe.name, "Максимум, м3": max_m3}
        for material in all_materials:
            value = required.get(material, Decimal("0"))
            row[f"Нужно, кг {material}"] = value
        rows.append(row)

    return pd.DataFrame(rows)


def _build_prices_dataframe(
    recipes: list[Recipe], balances: dict[str, float], prices: dict[str, dict[str, float]]
) -> pd.DataFrame:
    rows = []
    for recipe in recipes:
        max_m3, _ = calculate_max_cubic_meters(recipe, balances)
        price = prices.get(_normalize_name(recipe.name), {})
        price_no_delivery_no_vat = price.get("no_delivery_no_vat", 0.0)
        price_no_delivery_vat = price.get("no_delivery_vat_22", 0.0)
        price_pickup_no_vat = price.get("pickup_no_vat", 0.0)
        price_pickup_vat = price.get("pickup_vat_22", 0.0)

        row = {
            "Наименование": recipe.name,
            "Стоимость без доставки без НДС": _money(max_m3, price_no_delivery_no_vat),
            "Стоимость без доставки с НДС 22%": _money(max_m3, price_no_delivery_vat),
            "Стоимость самовывоз без НДС": _money(max_m3, price_pickup_no_vat),
            "Стоимость самовывоз с НДС 22%": _money(max_m3, price_pickup_vat),
            " ": "",
            "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС": price_no_delivery_no_vat,
            "БЕЗ ДОСТАВКИ С НДС 22%": price_no_delivery_vat,
            "САМОВЫВОЗ БЕЗ НДС": price_pickup_no_vat,
            "ОКРУГЛ. САМОВЫВОЗ С НДС 22%": price_pickup_vat,
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    BLANK_LEFT = "__blank_left__"
    BLANK_COST_1 = "__blank_cost_1__"
    BLANK_COST_2 = "__blank_cost_2__"
    BLANK_ROUND_1 = "__blank_round_1__"
    BLANK_ROUND_2 = "__blank_round_2__"

    df[BLANK_LEFT] = ""
    df[BLANK_COST_1] = ""
    df[BLANK_COST_2] = ""
    df[BLANK_ROUND_1] = ""
    df[BLANK_ROUND_2] = ""

    name_col = "Наименование"
    c1 = "Стоимость без доставки без НДС"
    c2 = "Стоимость без доставки с НДС 22%"
    c3 = "Стоимость самовывоз без НДС"
    c4 = "Стоимость самовывоз с НДС 22%"
    spacer = " "
    r1 = "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС"
    r2 = "БЕЗ ДОСТАВКИ С НДС 22%"
    r3 = "САМОВЫВОЗ БЕЗ НДС"
    r4 = "ОКРУГЛ. САМОВЫВОЗ С НДС 22%"

    ordered = [
        name_col,
        BLANK_LEFT,
        c1,
        c2,
        BLANK_COST_1,
        c3,
        c4,
        spacer,
        r1,
        r2,
        BLANK_ROUND_1,
        BLANK_ROUND_2,
        r3,
        r4,
    ]

    df = df[ordered]
    return df


async def handle_start(message: Message) -> None:
    await message.answer(
        "Пришлите Excel с остатками. Я верну Excel с максимумом м3 по каждому типу."
    )


async def handle_document(message: Message, bot: Bot) -> None:
    if not message.document:
        return
    filename = message.document.file_name or "остатки.xlsx"
    if not filename.lower().endswith(".xlsx"):
        await message.answer("Поддерживаются только файлы Excel .xlsx.")
        return

    await message.answer("Файл получен, считаю...")

    materials = _load_materials()
    recipes = _load_recipes()
    prices = _load_prices()

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        input_path = tmp_path / filename
        tg_file = await bot.get_file(message.document.file_id)
        await bot.download_file(tg_file.file_path, destination=input_path)

        try:
            balances = extract_balances(str(input_path), materials)
        except Exception as exc:
            await message.answer(f"Ошибка чтения файла: {exc}")
            return

        output_df = _build_output_dataframe(recipes, balances)
        prices_df = _build_prices_dataframe(recipes, balances, prices)
        output_path = tmp_path / "результат.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            output_df.to_excel(writer, index=False, startrow=0, sheet_name="Итог")
            start_row = len(output_df.index) + 5
            prices_df.to_excel(writer, index=False, startrow=start_row, sheet_name="Итог")

            ws = writer.book["Итог"]
            header_font = Font(bold=True)
            header_align = Alignment(
                horizontal="justify", vertical="center", wrap_text=True
            )
            body_align = Alignment(horizontal="justify", vertical="center")
            thin = Side(style="thin")
            table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            spacer_border = Border(left=thin, right=thin)
            palette = [
                "EAF2F8",
                "E8F6F3",
                "FEF9E7",
                "FDEDEC",
                "F4ECF7",
                "FDF2E9",
                "EBF5FB",
                "EAECEE",
            ]
            palette_bright = [
                "BBDFF7",
                "BFEED4",
                "F9E79F",
                "F5B7B1",
                "D7BDE2",
                "F8CFA8",
                "BBDFF7",
                "CCD1D1",
            ]
            highlight_b = PatternFill(
                start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
            )
            highlight_b_bright = PatternFill(
                start_color="A9D6F4", end_color="A9D6F4", fill_type="solid"
            )

            highlight_names = {
                _normalize_name(name)
                for name in [
                    "БСТ ВЗ0 F150 W6",
                    "БСТ В20 F150 W6",
                    "БСТ В25 F150 W8",
                    "БСТ B7,5 F50 W2",
                    "БСТ В20",
                    "БСТ В25 F150 W6",
                ]
            }

            output_header_row = 1
            output_start_row = output_header_row + 1
            output_end_row = output_start_row + len(output_df.index) - 1
            output_end_col = output_df.shape[1]

            prices_header_row = start_row + 1
            prices_start_row = prices_header_row + 1
            prices_end_row = prices_start_row + len(prices_df.index) - 1
            prices_end_col = prices_df.shape[1]

            for row_idx in range(output_header_row, output_end_row + 1):
                for cell in ws.iter_rows(
                    min_row=row_idx, max_row=row_idx, min_col=1, max_col=output_end_col
                ):
                    for c in cell:
                        is_highlight = False
                        if row_idx >= output_start_row:
                            data_idx = row_idx - output_start_row
                            if 0 <= data_idx < len(output_df.index):
                                name = output_df.iloc[data_idx, 0]
                                is_highlight = _normalize_name(str(name)) in highlight_names
                        if c.value is None or str(c.value).strip() == "":
                            c.border = spacer_border
                            c.fill = PatternFill(fill_type=None)
                        elif c.column == 2:
                            c.border = table_border
                            c.fill = highlight_b_bright if is_highlight else highlight_b
                        else:
                            c.border = table_border
                            fill_source = palette_bright if is_highlight else palette
                            fill_color = fill_source[(c.column - 1) % len(fill_source)]
                            c.fill = PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type="solid",
                            )
                        if row_idx == output_header_row:
                            c.font = header_font
                            c.alignment = header_align
                        else:
                            c.alignment = body_align
                            if isinstance(c.value, (int, float, Decimal)):
                                c.number_format = "#,##0.00"

            blank_cols = set()
            for blank_name in [
                "__blank_left__",
                "__blank_cost_1__",
                "__blank_cost_2__",
                "__blank_round_1__",
                "__blank_round_2__",
                " ",
            ]:
                if blank_name in prices_df.columns:
                    blank_cols.add(prices_df.columns.get_loc(blank_name) + 1)

            for row_idx in range(prices_header_row, prices_end_row + 1):
                for cell in ws.iter_rows(
                    min_row=row_idx, max_row=row_idx, min_col=1, max_col=prices_end_col
                ):
                    for c in cell:
                        is_highlight = False
                        if row_idx >= prices_start_row:
                            data_idx = row_idx - prices_start_row
                            if 0 <= data_idx < len(prices_df.index):
                                name = prices_df.iloc[data_idx, 0]
                                is_highlight = _normalize_name(str(name)) in highlight_names
                        if c.column in blank_cols:
                            c.border = Border()
                            c.fill = PatternFill(fill_type=None)
                        elif c.value is None or str(c.value).strip() == "":
                            c.border = spacer_border
                            c.fill = PatternFill(fill_type=None)
                        elif c.column == 2:
                            c.border = table_border
                            c.fill = highlight_b_bright if is_highlight else highlight_b
                        else:
                            c.border = table_border
                            fill_source = palette_bright if is_highlight else palette
                            fill_color = fill_source[(c.column - 1) % len(fill_source)]
                            c.fill = PatternFill(
                                start_color=fill_color,
                                end_color=fill_color,
                                fill_type="solid",
                            )
                        if row_idx == prices_header_row:
                            c.font = header_font
                            c.alignment = header_align
                        else:
                            c.alignment = body_align
                            if isinstance(c.value, (int, float, Decimal)):
                                c.number_format = "#,##0.00"

            price_columns = set()
            for idx, name in enumerate(prices_df.columns, start=1):
                if "Стоимость" in name or "БЕЗ ДОСТАВКИ" in name or "САМОВЫВОЗ" in name:
                    if name.strip():
                        price_columns.add(idx)

            # выделяем максимальные значения по каждой ценовой колонке жирным
            price_cols = [
                "Стоимость без доставки без НДС",
                "Стоимость без доставки с НДС 22%",
                "Стоимость самовывоз без НДС",
                "Стоимость самовывоз с НДС 22%",
            ]
            for col_name in price_cols:
                if col_name not in prices_df.columns:
                    continue
                col_idx = prices_df.columns.get_loc(col_name) + 1
                try:
                    series = prices_df[col_name]
                    max_val = series.max()
                except Exception:
                    continue
                if pd.isna(max_val):
                    continue
                for row_offset, val in enumerate(series, start=0):
                    if pd.isna(val):
                        continue
                    if abs(Decimal(str(val)) - Decimal(str(max_val))) > Decimal(
                        "0.0000001"
                    ):
                        continue
                    excel_row = prices_start_row + row_offset
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.font = Font(bold=True)

            # очищаем заголовки у служебных пустых колонок
            for blank_name in [
                "__blank_left__",
                "__blank_cost_1__",
                "__blank_cost_2__",
                "__blank_round_1__",
                "__blank_round_2__",
            ]:
                if blank_name in prices_df.columns:
                    b_col = prices_df.columns.get_loc(blank_name) + 1
                    ws.cell(row=prices_header_row, column=b_col).value = ""

            # группирующие заголовки над блоками цен
            group_row = prices_header_row - 1
            # стоимости
            ws.merge_cells(
                start_row=group_row, start_column=3, end_row=group_row, end_column=4
            )
            g1 = ws.cell(row=group_row, column=3, value="для организации А")
            ws.merge_cells(
                start_row=group_row, start_column=6, end_row=group_row, end_column=7
            )
            g2 = ws.cell(row=group_row, column=6, value="для иных организаций")
            # округлённые цены
            ws.merge_cells(
                start_row=group_row, start_column=9, end_row=group_row, end_column=10
            )
            g3 = ws.cell(row=group_row, column=9, value="для организации А")
            ws.merge_cells(
                start_row=group_row, start_column=13, end_row=group_row, end_column=14
            )
            g4 = ws.cell(row=group_row, column=13, value="для иных организаций")

            title_font_size = (header_font.sz or 11) + 1
            for gcell in (g1, g2, g3, g4):
                gcell.font = Font(bold=True, size=title_font_size)
                gcell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            for column_cells in ws.columns:
                max_len = 0
                col = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value is None:
                        continue
                    max_len = max(max_len, len(str(cell.value)))
                if max_len:
                    auto_width = min(max_len + 2, 60)
                    ws.column_dimensions[col].width = max(auto_width, 10)

        document = FSInputFile(output_path)
        await message.answer_document(document=document, filename="результат.xlsx")


async def main() -> None:
    bot = Bot(token=get_bot_token())
    dp = Dispatcher()
    dp.message.register(handle_start, CommandStart())
    dp.message.register(handle_document, F.document)

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
