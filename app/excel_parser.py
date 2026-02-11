from __future__ import annotations

from dataclasses import dataclass
import re
from typing import Iterable

from openpyxl import load_workbook


def _normalize(text: str) -> str:
    text = (
        text.strip()
        .lower()
        .replace("ё", "е")
        .replace("\xa0", " ")
        .replace("–", "-")
    )
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_match(text: str) -> str:
    text = _normalize(text)
    text = re.sub(r"\((т|кг|м3|м³)\)", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _parse_quantity(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("\xa0", "").replace(" ", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _find_saldo_column(ws) -> int | None:
    target = "сальдо на конец периода"
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and target in _normalize(cell.value):
                return cell.column
    return None


def _find_rows_for_alias(ws, alias: str, columns: Iterable[int]) -> list[tuple[int, bool]]:
    target = _normalize_match(alias)
    matches: list[tuple[int, bool]] = []
    for col in columns:
        for row in ws.iter_rows(min_col=col, max_col=col):
            cell = row[0]
            if not isinstance(cell.value, str):
                continue
            cell_norm = _normalize_match(cell.value)
            if not cell_norm:
                continue
            if cell_norm == target:
                matches.append((cell.row, True))
            elif target in cell_norm:
                matches.append((cell.row, False))
    return matches


@dataclass(frozen=True)
class MaterialConfig:
    name: str
    aliases: list[str]


def extract_balances(file_path: str, materials: list[MaterialConfig]) -> dict[str, float]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    saldo_col = _find_saldo_column(ws)
    if saldo_col is None:
        raise ValueError("Не найден столбец 'Сальдо на конец периода'")

    balances: dict[str, float] = {}
    for material in materials:
        matches: list[tuple[int, bool]] = []
        for alias in material.aliases:
            matches.extend(_find_rows_for_alias(ws, alias, columns=[1, 2]))

        seen_rows: set[int] = set()
        unique_matches = []
        for row, is_exact in matches:
            if row in seen_rows:
                continue
            seen_rows.add(row)
            unique_matches.append((row, is_exact))

        best_value = 0.0
        best_exact = False
        for row, is_exact in unique_matches:
            value = _parse_quantity(ws.cell(row=row + 1, column=saldo_col).value)
            if value == 0.0:
                continue
            if is_exact and not best_exact:
                best_exact = True
                best_value = value
                continue
            if is_exact == best_exact and value > best_value:
                best_value = value

        if best_value == 0.0 and unique_matches:
            row, _ = unique_matches[0]
            best_value = _parse_quantity(ws.cell(row=row + 1, column=saldo_col).value)

        balances[material.name] = best_value
    return balances
