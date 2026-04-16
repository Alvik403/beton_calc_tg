from __future__ import annotations

import io
import json
import re
import tempfile
import time
from decimal import Decimal, ROUND_FLOOR
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from fastapi import Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse

from app.calculator import Recipe, calculate_max_cubic_meters, calculate_recipe_diagnostics
from app.config import (
    load_materials_config,
    load_prices_config,
    load_recipes_config,
)
from app.directions import get_all_directions, get_direction, validate_scope
from app.excel_parser import MaterialConfig, extract_balances


app = FastAPI(title="Бетон калькулятор")


RATE_LIMIT_SECONDS = 10
_last_request_per_ip: dict[str, float] = {}

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_DIR = BASE_DIR / "config"
CONFIG_PASSWORD = "06032026"


def _profiles_path(scope: Optional[str]) -> Path:
    return get_direction(validate_scope(scope)).profiles_path


def _default_config_payload(scope: Optional[str]) -> Dict[str, Any]:
    return get_direction(validate_scope(scope)).get_default_config()


def _build_left_stack_html() -> str:
    """Генерирует левую панель из реестра направлений."""
    directions = get_all_directions()
    parts = []
    legacy_ids = [
        ("cfgBtn", "calcForm", "mainProfileSelect", "file", "calcOnlyBtn", "downloadBtn"),
        ("jbiCfgBtn", "jbiForm", "jbiProfileSelect", "jbiFile", "jbiCalcOnlyBtn", "jbiDownloadBtn"),
    ]
    for i, d in enumerate(directions):
        cfg_id, form_id, profile_id, file_id, calc_id, dl_id = (
            legacy_ids[i] if i < len(legacy_ids) else
            (f"cfgBtn-{d.id}", f"form-{d.id}", f"profile-{d.id}", f"file-{d.id}", f"calcBtn-{d.id}", f"downloadBtn-{d.id}")
        )
        dl_style = ' style="display:none"' if not d.supports_excel else ""
        parts.append(
            f'''
                        <div class="stack-section" data-direction-id="{d.id}">
                            <div class="rail-section-title">{d.display_name}</div>
                            <div class="box">
                                <button type="button" class="cfg-btn" id="{cfg_id}" title="Настройки" data-scope="{d.id}">⚙</button>
                                <h1>Загрузка файла</h1>
                                <form id="{form_id}" method="post" action="/upload" enctype="multipart/form-data" data-scope="{d.id}">
                                    <input type="hidden" name="scope" value="{d.id}" />
                                    <div class="field">
                                        <label for="{profile_id}">Считать по настройкам</label>
                                        <select id="{profile_id}" class="main-profile-select" name="profile_name">
                                            <option value="__base__">По умолчанию</option>
                                        </select>
                                    </div>
                                    <div class="field">
                                        <label for="{file_id}">Файл .xlsx</label>
                                        <input id="{file_id}" name="file" type="file" accept=".xlsx" required />
                                    </div>
                                    <div class="hp-field">
                                        <label>Ваш сайт</label>
                                        <input type="text" name="website" autocomplete="off" />
                                    </div>
                                    <div class="btn-row">
                                        <button class="btn" type="button" id="{calc_id}">Посчитать</button>
                                        <button class="btn" type="button" id="{dl_id}"{dl_style}>Скачать</button>
                                    </div>
                                </form>
                            </div>
                        </div>'''
        )
    return "\n".join(parts)


def _require_config_password(request: Request) -> None:
    if request.headers.get("X-Config-Password") != CONFIG_PASSWORD:
        raise HTTPException(status_code=401, detail="Неверный пароль конфигуратора")


def _load_profiles(scope: Optional[str] = None) -> Dict[str, Any]:
    path = _profiles_path(scope)
    if not path.exists():
        return {"profiles": [], "active": None}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
            if not isinstance(data, dict):
                return {"profiles": [], "active": None}
            data.setdefault("profiles", [])
            data.setdefault("active", None)
            return data
    except Exception:
        return {"profiles": [], "active": None}


def _save_profiles(data: Dict[str, Any], scope: Optional[str] = None) -> None:
    path = _profiles_path(scope)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _get_profile(data: Dict[str, Any], name: Optional[str]) -> Optional[Dict[str, Any]]:
    if not name:
        return None
    for p in data.get("profiles", []):
        if p.get("name") == name:
            return p
    return None


def _client_ip(request: Request) -> str:
    client = request.client
    return client.host if client else "unknown"


def _is_rate_limited(ip: str) -> bool:
    now = time.time()
    last = _last_request_per_ip.get(ip)
    if last is not None and now - last < RATE_LIMIT_SECONDS:
        return True
    _last_request_per_ip[ip] = now
    return False


def _money(volume_m3: Decimal, price_per_m3: float) -> Decimal:
    return Decimal(str(volume_m3)) * Decimal(str(price_per_m3))


def _normalize_name(text: str) -> str:
    text = text.strip().lower().replace("ё", "е")
    text = text.replace("в", "b").replace("з", "3")
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_alias_for_validation(text: str) -> str:
    text = text.strip().lower().replace("ё", "е").replace("\xa0", " ")
    text = text.replace("–", "-")
    text = re.sub(r"\((т|кг|м3|м³)\)", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError("Ожидалось числовое значение") from exc


def _validation_error(field: str, message: str, index: Optional[int] = None) -> Dict[str, Any]:
    error = {"field": field, "message": message}
    if index is not None:
        error["index"] = index
    return error


def _raise_validation_error(errors: List[Dict[str, Any]]) -> None:
    raise HTTPException(
        status_code=400,
        detail={
            "message": "Ошибка валидации конфигурации",
            "errors": errors,
        },
    )


def _validate_and_prepare_profile(
    payload: Dict[str, Any], scope: Optional[str] = None
) -> Dict[str, Any]:
    scope = validate_scope(scope)
    errors: List[Dict[str, Any]] = []
    name = str(payload.get("name") or "").strip()
    if not name:
        errors.append(_validation_error("name", "Имя профиля обязательно"))
    elif name == "__base__":
        errors.append(
            _validation_error("name", 'Имя "__base__" зарезервировано и недоступно для профиля')
        )

    raw_materials = payload.get("materials") or []
    raw_recipes = payload.get("recipes") or []
    raw_prices = payload.get("prices") or []

    if not isinstance(raw_materials, list):
        errors.append(_validation_error("materials", "Материалы должны быть списком"))
        raw_materials = []
    if not isinstance(raw_recipes, list):
        errors.append(_validation_error("recipes", "Составы должны быть списком"))
        raw_recipes = []
    if not isinstance(raw_prices, list):
        errors.append(_validation_error("prices", "Цены должны быть списком"))
        raw_prices = []

    material_names_seen: dict[str, int] = {}
    alias_to_material: dict[str, str] = {}
    materials: List[Dict[str, Any]] = []

    for idx, item in enumerate(raw_materials):
        if not isinstance(item, dict):
            errors.append(_validation_error("materials", "Материал должен быть объектом", idx))
            continue
        material_name = str(item.get("name") or "").strip()
        raw_aliases = item.get("aliases") or []
        if not material_name:
            errors.append(
                _validation_error("materials", "У материала должно быть наименование", idx)
            )
        normalized_material_name = _normalize_name(material_name) if material_name else ""
        if normalized_material_name:
            prev_idx = material_names_seen.get(normalized_material_name)
            if prev_idx is not None:
                errors.append(
                    _validation_error(
                        "materials",
                        f'Дублируется материал "{material_name}"',
                        idx,
                    )
                )
            else:
                material_names_seen[normalized_material_name] = idx

        if not isinstance(raw_aliases, list):
            errors.append(
                _validation_error("materials", "Список алиасов должен быть массивом", idx)
            )
            raw_aliases = []

        aliases: List[str] = []
        aliases_seen: set[str] = set()
        for alias in raw_aliases:
            alias_text = str(alias or "").strip()
            if not alias_text:
                continue
            normalized_alias = _normalize_alias_for_validation(alias_text)
            if not normalized_alias or normalized_alias in aliases_seen:
                continue
            aliases_seen.add(normalized_alias)
            aliases.append(alias_text)
            owner = alias_to_material.get(normalized_alias)
            if owner is not None and owner != material_name:
                errors.append(
                    _validation_error(
                        "materials",
                        f'Алиас "{alias_text}" конфликтует между "{owner}" и "{material_name}"',
                        idx,
                    )
                )
            elif material_name:
                alias_to_material[normalized_alias] = material_name

        if not aliases:
            errors.append(
                _validation_error(
                    "materials",
                    f'Для материала "{material_name or "без названия"}" нужен хотя бы один алиас',
                    idx,
                )
            )

        materials.append({"name": material_name, "aliases": aliases})

    available_material_names = {item["name"] for item in materials if item.get("name")}
    if scope == "jbi":
        available_material_names.update(
            recipe.name for recipe in _load_recipes(scope="beton")
        )

    recipe_names_seen: dict[str, int] = {}
    recipes: List[Dict[str, Any]] = []
    for idx, item in enumerate(raw_recipes):
        if not isinstance(item, dict):
            errors.append(_validation_error("recipes", "Состав должен быть объектом", idx))
            continue
        recipe_name = str(item.get("name") or "").strip()
        recipe_materials = item.get("materials") or {}
        if not recipe_name:
            errors.append(
                _validation_error("recipes", "У состава должно быть наименование", idx)
            )
        normalized_recipe_name = _normalize_name(recipe_name) if recipe_name else ""
        if normalized_recipe_name:
            prev_idx = recipe_names_seen.get(normalized_recipe_name)
            if prev_idx is not None:
                errors.append(
                    _validation_error(
                        "recipes",
                        f'Дублируется состав "{recipe_name}"',
                        idx,
                    )
                )
            else:
                recipe_names_seen[normalized_recipe_name] = idx
        if not isinstance(recipe_materials, dict):
            errors.append(
                _validation_error("recipes", "Материалы состава должны быть объектом", idx)
            )
            recipe_materials = {}

        sanitized_recipe_materials: Dict[str, float] = {}
        positive_count = 0
        for material_name, raw_amount in recipe_materials.items():
            mat_name = str(material_name or "").strip()
            if not mat_name:
                errors.append(
                    _validation_error("recipes", "В составе найден пустой материал", idx)
                )
                continue
            if mat_name not in available_material_names:
                errors.append(
                    _validation_error(
                        "recipes",
                        f'Состав "{recipe_name or "без названия"}" ссылается на неизвестный материал "{mat_name}"',
                        idx,
                    )
                )
            try:
                amount = _to_float(raw_amount)
            except ValueError:
                errors.append(
                    _validation_error(
                        "recipes",
                        f'Для материала "{mat_name}" в составе "{recipe_name or "без названия"}" указано нечисловое значение',
                        idx,
                    )
                )
                continue
            if amount < 0:
                errors.append(
                    _validation_error(
                        "recipes",
                        f'Для материала "{mat_name}" в составе "{recipe_name or "без названия"}" нельзя указывать отрицательное значение',
                        idx,
                    )
                )
            if amount > 0:
                positive_count += 1
            sanitized_recipe_materials[mat_name] = amount

        if not sanitized_recipe_materials or positive_count == 0:
            errors.append(
                _validation_error(
                    "recipes",
                    f'Состав "{recipe_name or "без названия"}" должен содержать хотя бы один материал с расходом больше нуля',
                    idx,
                )
            )

        recipes.append({"name": recipe_name, "materials": sanitized_recipe_materials})

    price_names_seen: dict[str, int] = {}
    price_fields = (
        "no_delivery_no_vat",
        "no_delivery_vat_22",
        "pickup_no_vat",
        "pickup_vat_22",
    )
    prices: List[Dict[str, Any]] = []
    for idx, item in enumerate(raw_prices):
        if not isinstance(item, dict):
            errors.append(_validation_error("prices", "Цена должна быть объектом", idx))
            continue
        price_name = str(item.get("name") or "").strip()
        if not price_name:
            errors.append(_validation_error("prices", "У цены должно быть наименование", idx))
        normalized_price_name = _normalize_name(price_name) if price_name else ""
        if normalized_price_name:
            prev_idx = price_names_seen.get(normalized_price_name)
            if prev_idx is not None:
                errors.append(
                    _validation_error(
                        "prices",
                        f'Дублируется цена для "{price_name}"',
                        idx,
                    )
                )
            else:
                price_names_seen[normalized_price_name] = idx

        sanitized_price: Dict[str, Any] = {"name": price_name}
        for field in price_fields:
            try:
                value = _to_float(item.get(field, 0) or 0)
            except ValueError:
                errors.append(
                    _validation_error(
                        "prices",
                        f'Поле "{field}" для "{price_name or "без названия"}" должно быть числом',
                        idx,
                    )
                )
                value = 0.0
            if value < 0:
                errors.append(
                    _validation_error(
                        "prices",
                        f'Поле "{field}" для "{price_name or "без названия"}" не может быть отрицательным',
                        idx,
                    )
                )
            sanitized_price[field] = value
        prices.append(sanitized_price)

    if errors:
        _raise_validation_error(errors)

    return {
        "name": name,
        "materials": materials,
        "recipes": recipes,
        "prices": prices,
    }


def _load_materials(
    scope: Optional[str] = None, profile_name: Optional[str] = None
) -> list[MaterialConfig]:
    scope = validate_scope(scope)
    profiles = _load_profiles(scope)
    active_name = profile_name if profile_name is not None else profiles.get("active")
    active = _get_profile(profiles, active_name)
    base = _default_config_payload(scope)["materials"]
    raw = active.get("materials") if active and "materials" in active else base
    materials: list[MaterialConfig] = []
    for item in raw:
        name = item.get("name", "").strip()
        aliases = [a for a in item.get("aliases", []) if a]
        if name and aliases:
            materials.append(MaterialConfig(name=name, aliases=aliases))
    return materials


def _load_recipes(
    scope: Optional[str] = None, profile_name: Optional[str] = None
) -> list[Recipe]:
    scope = validate_scope(scope)
    profiles = _load_profiles(scope)
    active_name = profile_name if profile_name is not None else profiles.get("active")
    active = _get_profile(profiles, active_name)
    base = _default_config_payload(scope)["recipes"]
    raw = active.get("recipes") if active and "recipes" in active else base
    recipes: list[Recipe] = []
    for item in raw:
        name = item.get("name", "").strip()
        materials = item.get("materials", {})
        if name and materials:
            recipes.append(Recipe(name=name, materials=materials))
    return recipes


def _load_prices(
    scope: Optional[str] = None, profile_name: Optional[str] = None
) -> dict[str, dict[str, float]]:
    scope = validate_scope(scope)
    profiles = _load_profiles(scope)
    active_name = profile_name if profile_name is not None else profiles.get("active")
    active = _get_profile(profiles, active_name)
    raw_list: List[Dict[str, Any]]
    if active and "prices" in active:
        raw_list = active.get("prices") or []
    else:
        raw_list = _default_config_payload(scope)["prices"]
    prices: dict[str, dict[str, float]] = {}
    for item in raw_list:
        name = item.get("name", "").strip()
        if not name:
            continue
        prices[_normalize_name(name)] = {
            "no_delivery_no_vat": float(item.get("no_delivery_no_vat", 0) or 0),
            "no_delivery_vat_22": float(item.get("no_delivery_vat_22", 0) or 0),
            "pickup_no_vat": float(item.get("pickup_no_vat", 0) or 0),
            "pickup_vat_22": float(item.get("pickup_vat_22", 0) or 0),
        }
    return prices


def _build_summary_volume_df(
    recipes: list[Recipe], balances: dict[str, float]
) -> pd.DataFrame:
    """Одна строка на рецепт — для сводки на сайте и merge с ценами."""
    rows = []
    for recipe in recipes:
        max_m3, _ = calculate_max_cubic_meters(recipe, balances)
        rows.append({"Наименование": recipe.name, "Максимум, м3": max_m3})
    return pd.DataFrame(rows)


def _build_excel_first_table_beton(
    recipes: list[Recipe], balances: dict[str, float]
) -> tuple[pd.DataFrame, list[tuple[int, int]]]:
    """Первая таблица Excel: материалы по строкам, колонки 1–2 объединяются по рецепту."""
    rows: list[dict[str, Any]] = []
    merge_ranges: list[tuple[int, int]] = []
    for recipe in recipes:
        max_m3, required = calculate_max_cubic_meters(recipe, balances)
        mat_rows: list[tuple[str, Decimal]] = []
        for material in sorted(recipe.materials.keys()):
            val = required.get(material, Decimal("0"))
            if val != 0:
                mat_rows.append((material, val))
        if not mat_rows:
            mat_rows = [("—", Decimal("0"))]
        start_idx = len(rows)
        for i, (mat, val) in enumerate(mat_rows):
            rows.append(
                {
                    "Наименование": recipe.name if i == 0 else "",
                    "Максимум, м3": max_m3 if i == 0 else "",
                    "Материал": mat,
                    "Расход, кг": val,
                }
            )
        merge_ranges.append((start_idx, len(rows) - 1))
    return pd.DataFrame(rows), merge_ranges


def _build_excel_first_table_jbi(
    jbi_recipes: list[Recipe], effective_balances: dict[str, float]
) -> tuple[pd.DataFrame, list[tuple[int, int]]]:
    rows: list[dict[str, Any]] = []
    merge_ranges: list[tuple[int, int]] = []
    for recipe in jbi_recipes:
        max_units_raw, required, _ = calculate_recipe_diagnostics(recipe, effective_balances)
        max_units = int(max_units_raw.to_integral_value(rounding=ROUND_FLOOR))
        mat_rows: list[tuple[str, Decimal]] = []
        for material in sorted(recipe.materials.keys()):
            val = required.get(material, Decimal("0"))
            if val != 0:
                mat_rows.append((material, val))
        if not mat_rows:
            mat_rows = [("—", Decimal("0"))]
        start_idx = len(rows)
        for i, (mat, val) in enumerate(mat_rows):
            rows.append(
                {
                    "Наименование": recipe.name if i == 0 else "",
                    "Максимум, шт": max_units if i == 0 else "",
                    "Материал": mat,
                    "Расход, кг": val,
                }
            )
        merge_ranges.append((start_idx, len(rows) - 1))
    return pd.DataFrame(rows), merge_ranges


def _build_prices_dataframe(
    recipes: list[Recipe], balances: dict[str, float], prices: dict[str, dict[str, float]]
) -> pd.DataFrame:
    rows = []
    for recipe in recipes:
        max_m3, _ = calculate_max_cubic_meters(recipe, balances)
        price = prices.get(_normalize_name(recipe.name), {})
        price_no_delivery_no_vat = price.get("no_delivery_no_vat", 0.0)
        price_no_delivery_vat = price.get("no_delivery_vat_22", 0.0)
        price_pickup_no_vat = price.get("pickup_no_vat", 0.0)
        price_pickup_vat = price.get("pickup_vat_22", 0.0)

        row = {
            "Наименование": recipe.name,
            "Стоимость без доставки без НДС": _money(max_m3, price_no_delivery_no_vat),
            "Стоимость без доставки с НДС 22%": _money(max_m3, price_no_delivery_vat),
            "Стоимость самовывоз без НДС": _money(max_m3, price_pickup_no_vat),
            "Стоимость самовывоз с НДС 22%": _money(max_m3, price_pickup_vat),
            " ": "",
            "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС": price_no_delivery_no_vat,
            "БЕЗ ДОСТАВКИ С НДС 22%": price_no_delivery_vat,
            "САМОВЫВОЗ БЕЗ НДС": price_pickup_no_vat,
            "ОКРУГЛ. САМОВЫВОЗ С НДС 22%": price_pickup_vat,
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # добавляем по 1 пустой колонке между группами цен и отступ в 1 колонку от левого края
    BLANK_LEFT = "__blank_left__"
    BLANK_COST_1 = "__blank_cost_1__"
    BLANK_COST_2 = "__blank_cost_2__"
    BLANK_ROUND_1 = "__blank_round_1__"
    BLANK_ROUND_2 = "__blank_round_2__"

    df[BLANK_LEFT] = ""
    df[BLANK_COST_1] = ""
    df[BLANK_COST_2] = ""
    df[BLANK_ROUND_1] = ""
    df[BLANK_ROUND_2] = ""

    name_col = "Наименование"
    c1 = "Стоимость без доставки без НДС"
    c2 = "Стоимость без доставки с НДС 22%"
    c3 = "Стоимость самовывоз без НДС"
    c4 = "Стоимость самовывоз с НДС 22%"
    spacer = " "
    r1 = "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС"
    r2 = "БЕЗ ДОСТАВКИ С НДС 22%"
    r3 = "САМОВЫВОЗ БЕЗ НДС"
    r4 = "ОКРУГЛ. САМОВЫВОЗ С НДС 22%"

    ordered = [
        name_col,
        BLANK_LEFT,
        c1,
        c2,
        BLANK_COST_1,
        c3,
        c4,
        spacer,
        r1,
        r2,
        BLANK_ROUND_1,
        BLANK_ROUND_2,
        r3,
        r4,
    ]

    df = df[ordered]
    return df


def _jbi_effective_balances(raw_balances: dict[str, float]) -> dict[str, float]:
    """Остатки ЖБИ + лимиты по бетону (м³ по рецептам), как в _build_jbi_summary."""
    beton_materials = _load_materials(scope="beton")
    beton_recipes = _load_recipes(scope="beton")
    beton_balances = {
        key: value for key, value in raw_balances.items() if key in {m.name for m in beton_materials}
    }
    concrete_limits: dict[str, float] = {}
    for recipe in beton_recipes:
        max_m3, _ = calculate_max_cubic_meters(recipe, beton_balances)
        concrete_limits[recipe.name] = float(max_m3)
    effective = dict(raw_balances)
    effective.update(concrete_limits)
    return effective


def _build_jbi_prices_dataframe(
    jbi_recipes: list[Recipe],
    effective_balances: dict[str, float],
    prices: dict[str, dict[str, float]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for recipe in jbi_recipes:
        max_units_raw, _, _ = calculate_recipe_diagnostics(recipe, effective_balances)
        max_units = int(max_units_raw.to_integral_value(rounding=ROUND_FLOOR))
        price = prices.get(_normalize_name(recipe.name), {})
        pdu = float(price.get("no_delivery_no_vat", 0.0) or 0.0)
        pdv = float(price.get("no_delivery_vat_22", 0.0) or 0.0)
        ppu = float(price.get("pickup_no_vat", 0.0) or 0.0)
        ppv = float(price.get("pickup_vat_22", 0.0) or 0.0)
        mu = Decimal(str(max_units))
        row = {
            "Наименование": recipe.name,
            "Стоимость без доставки без НДС": _money(mu, pdu),
            "Стоимость без доставки с НДС 22%": _money(mu, pdv),
            "Стоимость самовывоз без НДС": _money(mu, ppu),
            "Стоимость самовывоз с НДС 22%": _money(mu, ppv),
            " ": "",
            "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС": pdu,
            "БЕЗ ДОСТАВКИ С НДС 22%": pdv,
            "САМОВЫВОЗ БЕЗ НДС": ppu,
            "ОКРУГЛ. САМОВЫВОЗ С НДС 22%": ppv,
        }
        rows.append(row)
    df = pd.DataFrame(rows)
    BLANK_LEFT = "__blank_left__"
    BLANK_COST_1 = "__blank_cost_1__"
    BLANK_COST_2 = "__blank_cost_2__"
    BLANK_ROUND_1 = "__blank_round_1__"
    BLANK_ROUND_2 = "__blank_round_2__"
    df[BLANK_LEFT] = ""
    df[BLANK_COST_1] = ""
    df[BLANK_COST_2] = ""
    df[BLANK_ROUND_1] = ""
    df[BLANK_ROUND_2] = ""
    name_col = "Наименование"
    c1 = "Стоимость без доставки без НДС"
    c2 = "Стоимость без доставки с НДС 22%"
    c3 = "Стоимость самовывоз без НДС"
    c4 = "Стоимость самовывоз с НДС 22%"
    spacer = " "
    r1 = "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС"
    r2 = "БЕЗ ДОСТАВКИ С НДС 22%"
    r3 = "САМОВЫВОЗ БЕЗ НДС"
    r4 = "ОКРУГЛ. САМОВЫВОЗ С НДС 22%"
    ordered = [
        name_col,
        BLANK_LEFT,
        c1,
        c2,
        BLANK_COST_1,
        c3,
        c4,
        spacer,
        r1,
        r2,
        BLANK_ROUND_1,
        BLANK_ROUND_2,
        r3,
        r4,
    ]
    return df[ordered]


def _workbook_bytes_from_tables(
    output_df: pd.DataFrame,
    prices_df: pd.DataFrame,
    *,
    merge_first_section: Optional[list[tuple[int, int]]] = None,
) -> bytes:
    """Общая разметка листа «Итог» для бетона и ЖБИ. Первая таблица: колонки 1–2 сгруппированы по рецепту."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, startrow=0, sheet_name="Итог")
        start_row = len(output_df.index) + 5
        prices_df.to_excel(writer, index=False, startrow=start_row, sheet_name="Итог")

        ws = writer.book["Итог"]
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True)
        header_align = Alignment(horizontal="justify", vertical="center", wrap_text=True)
        body_align = Alignment(horizontal="justify", vertical="center")
        thin = Side(style="thin")
        table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        spacer_border = Border(left=thin, right=thin)
        palette = [
            "EAF2F8",
            "E8F6F3",
            "FEF9E7",
            "FDEDEC",
            "F4ECF7",
            "FDF2E9",
            "EBF5FB",
            "EAECEE",
        ]
        palette_bright = [
            "BBDFF7",
            "BFEED4",
            "F9E79F",
            "F5B7B1",
            "D7BDE2",
            "F8CFA8",
            "BBDFF7",
            "CCD1D1",
        ]
        highlight_b = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        highlight_b_bright = PatternFill(
            start_color="A9D6F4", end_color="A9D6F4", fill_type="solid"
        )

        highlight_names = {
            _normalize_name(name)
            for name in [
                "БСТ ВЗ0 F150 W6",
                "БСТ В20 F150 W6",
                "БСТ В25 F150 W8",
                "БСТ B7,5 F50 W2",
                "БСТ В20",
                "БСТ В25 F150 W6",
            ]
        }

        output_header_row = 1
        output_start_row = output_header_row + 1
        output_end_row = output_start_row + len(output_df.index) - 1
        output_end_col = output_df.shape[1]

        merges = merge_first_section or []
        for r0, r1 in merges:
            if r1 > r0:
                er1, er2 = r0 + 2, r1 + 2
                ws.merge_cells(start_row=er1, start_column=1, end_row=er2, end_column=1)
                ws.merge_cells(start_row=er1, start_column=2, end_row=er2, end_column=2)

        prices_header_row = start_row + 1
        prices_start_row = prices_header_row + 1
        prices_end_row = prices_start_row + len(prices_df.index) - 1
        prices_end_col = prices_df.shape[1]

        # Заголовок первой таблицы
        for cell in ws.iter_rows(
            min_row=output_header_row, max_row=output_header_row, min_col=1, max_col=output_end_col
        ):
            for c in cell:
                c.font = header_font
                c.alignment = header_align
                c.border = table_border
                c.fill = PatternFill(
                    start_color=palette[(c.column - 1) % len(palette)],
                    end_color=palette[(c.column - 1) % len(palette)],
                    fill_type="solid",
                )

        # Тело первой таблицы: сгруппированные A–B; материал и расход по строкам
        for r0, r1 in merges:
            name_val = str(output_df.iloc[r0, 0] or "")
            is_highlight = _normalize_name(name_val) in highlight_names
            er1, er2 = r0 + 2, r1 + 2
            ca = ws.cell(row=er1, column=1)
            ca.border = table_border
            ca.fill = PatternFill(
                start_color=palette_bright[0] if is_highlight else palette[0],
                end_color=palette_bright[0] if is_highlight else palette[0],
                fill_type="solid",
            )
            ca.alignment = Alignment(horizontal="justify", vertical="center", wrap_text=True)
            cb = ws.cell(row=er1, column=2)
            cb.border = table_border
            cb.fill = highlight_b_bright if is_highlight else highlight_b
            cb.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cb.value, (int, float, Decimal)):
                cb.number_format = "#,##0.00"

            for r in range(r0, r1 + 1):
                er = r + 2
                for col in (3, 4):
                    c = ws.cell(row=er, column=col)
                    c.border = table_border
                    fill_idx = col - 1
                    fill_source = palette_bright if is_highlight else palette
                    fc = fill_source[fill_idx % len(fill_source)]
                    c.fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
                    c.alignment = body_align
                    if isinstance(c.value, (int, float, Decimal)):
                        c.number_format = "#,##0.00"

        blank_cols = set()
        for blank_name in [
            "__blank_left__",
            "__blank_cost_1__",
            "__blank_cost_2__",
            "__blank_round_1__",
            "__blank_round_2__",
            " ",
        ]:
            if blank_name in prices_df.columns:
                blank_cols.add(prices_df.columns.get_loc(blank_name) + 1)

        for row_idx in range(prices_header_row, prices_end_row + 1):
            for cell in ws.iter_rows(
                min_row=row_idx, max_row=row_idx, min_col=1, max_col=prices_end_col
            ):
                for c in cell:
                    is_highlight = False
                    if row_idx >= prices_start_row:
                        data_idx = row_idx - prices_start_row
                        if 0 <= data_idx < len(prices_df.index):
                            name = prices_df.iloc[data_idx, 0]
                            is_highlight = _normalize_name(str(name)) in highlight_names
                    if c.column in blank_cols:
                        c.border = Border()
                        c.fill = PatternFill(fill_type=None)
                    elif c.value is None or str(c.value).strip() == "":
                        c.border = spacer_border
                        c.fill = PatternFill(fill_type=None)
                    elif c.column == 2:
                        c.border = table_border
                        c.fill = highlight_b_bright if is_highlight else highlight_b
                    else:
                        c.border = table_border
                        fill_source = palette_bright if is_highlight else palette
                        fill_color = fill_source[(c.column - 1) % len(fill_source)]
                        c.fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type="solid",
                        )
                    if row_idx == prices_header_row:
                        c.font = header_font
                        c.alignment = header_align
                    else:
                        c.alignment = body_align
                        if isinstance(c.value, (int, float, Decimal)):
                            c.number_format = "#,##0.00"

        price_cols = [
            "Стоимость без доставки без НДС",
            "Стоимость без доставки с НДС 22%",
            "Стоимость самовывоз без НДС",
            "Стоимость самовывоз с НДС 22%",
        ]
        for col_name in price_cols:
            if col_name not in prices_df.columns:
                continue
            col_idx = prices_df.columns.get_loc(col_name) + 1
            try:
                series = prices_df[col_name]
                max_val = series.max()
            except Exception:
                continue
            if pd.isna(max_val):
                continue
            for row_offset, val in enumerate(series, start=0):
                if pd.isna(val):
                    continue
                if abs(Decimal(str(val)) - Decimal(str(max_val))) > Decimal("0.0000001"):
                    continue
                excel_row = prices_start_row + row_offset
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.font = Font(bold=True)

        for blank_name in [
            "__blank_left__",
            "__blank_cost_1__",
            "__blank_cost_2__",
            "__blank_round_1__",
            "__blank_round_2__",
        ]:
            if blank_name in prices_df.columns:
                b_col = prices_df.columns.get_loc(blank_name) + 1
                ws.cell(row=prices_header_row, column=b_col).value = ""

        group_row = prices_header_row - 1
        ws.merge_cells(start_row=group_row, start_column=3, end_row=group_row, end_column=4)
        g1 = ws.cell(row=group_row, column=3, value="для организации А")
        ws.merge_cells(start_row=group_row, start_column=6, end_row=group_row, end_column=7)
        g2 = ws.cell(row=group_row, column=6, value="для иных организаций")
        ws.merge_cells(start_row=group_row, start_column=9, end_row=group_row, end_column=10)
        g3 = ws.cell(row=group_row, column=9, value="для организации А")
        ws.merge_cells(start_row=group_row, start_column=13, end_row=group_row, end_column=14)
        g4 = ws.cell(row=group_row, column=13, value="для иных организаций")

        title_font_size = (header_font.sz or 11) + 1
        for gcell in (g1, g2, g3, g4):
            gcell.font = Font(bold=True, size=title_font_size)
            gcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                v = str(cell.value)
                if "\n" in v:
                    for line in v.split("\n"):
                        max_len = max(max_len, len(line))
                else:
                    max_len = max(max_len, len(v))
            if col_idx == 3:
                auto_width = min(max(max_len + 2, 46), 58)
            else:
                auto_width = min(max_len + 2, 60)
            ws.column_dimensions[letter].width = max(auto_width, 10)

    output.seek(0)
    return output.read()


def _build_summary(
    balances: dict[str, float],
    scope: Optional[str] = None,
    profile_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Построить краткую сводку по объемам и ценам для вывода на сайт."""
    recipes = _load_recipes(scope=scope, profile_name=profile_name)
    prices = _load_prices(scope=scope, profile_name=profile_name)

    output_df = _build_summary_volume_df(recipes, balances)
    prices_df = _build_prices_dataframe(recipes, balances, prices)

    name_col = "Наименование"
    m_col = "Максимум, м3"
    c1 = "Стоимость без доставки без НДС"
    c2 = "Стоимость без доставки с НДС 22%"
    c3 = "Стоимость самовывоз без НДС"
    c4 = "Стоимость самовывоз с НДС 22%"
    r1 = "Округл. БЕЗ ДОСТАВКИ БЕЗ НДС"
    r2 = "БЕЗ ДОСТАВКИ С НДС 22%"
    r3 = "САМОВЫВОЗ БЕЗ НДС"
    r4 = "ОКРУГЛ. САМОВЫВОЗ С НДС 22%"

    merged = pd.merge(
        output_df[[name_col, m_col]],
        prices_df[[name_col, c1, c2, c3, c4, r1, r2, r3, r4]],
        on=name_col,
        how="left",
    )

    items: list[Dict[str, Any]] = []
    total_volume = Decimal("0")
    for _, row in merged.iterrows():
        name = str(row.get(name_col, "") or "")
        max_m3 = Decimal(str(row.get(m_col, 0) or 0))
        total_volume += max_m3
        recipe = next((item for item in recipes if item.name == name), None)
        limiter_data = []
        if recipe is not None:
            _, _, limiters = calculate_recipe_diagnostics(recipe, balances)
            limiter_data = [
                {
                    "material": limiter["material"],
                    "available": float(limiter["available"]),
                    "required_per_unit": float(limiter["required_per_unit"]),
                    "possible_output": float(limiter["possible_output"]),
                }
                for limiter in limiters
            ]

        def _val(col: str) -> Optional[float]:
            v = row.get(col, None)
            if pd.isna(v):
                return None
            try:
                return float(v)
            except Exception:
                return None

        items.append(
            {
                "name": name,
                "max_m3": float(max_m3),
                "amounts": {
                    "no_delivery_no_vat": _val(c1),
                    "no_delivery_vat_22": _val(c2),
                    "pickup_no_vat": _val(c3),
                    "pickup_vat_22": _val(c4),
                },
                "unit_prices": {
                    "no_delivery_no_vat": _val(r1),
                    "no_delivery_vat_22": _val(r2),
                    "pickup_no_vat": _val(r3),
                    "pickup_vat_22": _val(r4),
                },
                "limiters": limiter_data,
            }
        )

    return {
        "kind": "beton",
        "items": items,
        "total_volume": float(total_volume),
    }


def _build_jbi_summary(
    raw_balances: dict[str, float], profile_name: Optional[str] = None
) -> Dict[str, Any]:
    """Рассчитать максимум изделий ЖБИ с учетом доступного бетона."""
    jbi_recipes = _load_recipes(scope="jbi", profile_name=profile_name)
    jbi_prices = _load_prices(scope="jbi", profile_name=profile_name)

    effective_balances = _jbi_effective_balances(raw_balances)

    items: list[Dict[str, Any]] = []
    for recipe in jbi_recipes:
        max_units_raw, _, limiters = calculate_recipe_diagnostics(recipe, effective_balances)
        max_units = int(max_units_raw.to_integral_value(rounding=ROUND_FLOOR))
        price = jbi_prices.get(_normalize_name(recipe.name), {})
        unit_price = float(price.get("no_delivery_no_vat", 0.0) or 0.0)
        items.append(
            {
                "name": recipe.name,
                "max_units": max_units,
                "unit_price": unit_price,
                "total_price": float(Decimal(str(unit_price)) * Decimal(str(max_units))),
                "limiters": [
                    {
                        "material": limiter["material"],
                        "available": float(limiter["available"]),
                        "required_per_unit": float(limiter["required_per_unit"]),
                        "possible_output": float(limiter["possible_output"]),
                    }
                    for limiter in limiters
                ],
            }
        )

    return {
        "kind": "jbi",
        "items": items,
    }


def _build_jbi_excel(
    raw_balances: dict[str, float], profile_name: Optional[str] = None
) -> bytes:
    """Excel «Итог» для ЖБИ: максимум шт, расход материалов, четыре варианта цен (как у бетона)."""
    jbi_recipes = _load_recipes(scope="jbi", profile_name=profile_name)
    jbi_prices = _load_prices(scope="jbi", profile_name=profile_name)
    effective = _jbi_effective_balances(raw_balances)
    output_df, merge_ranges = _build_excel_first_table_jbi(jbi_recipes, effective)
    prices_df = _build_jbi_prices_dataframe(jbi_recipes, effective, jbi_prices)
    return _workbook_bytes_from_tables(output_df, prices_df, merge_first_section=merge_ranges)


def _build_excel(
    balances: dict[str, float],
    scope: Optional[str] = None,
    profile_name: Optional[str] = None,
) -> bytes:
    recipes = _load_recipes(scope=scope, profile_name=profile_name)
    prices = _load_prices(scope=scope, profile_name=profile_name)

    output_df, merge_ranges = _build_excel_first_table_beton(recipes, balances)
    prices_df = _build_prices_dataframe(recipes, balances, prices)
    return _workbook_bytes_from_tables(output_df, prices_df, merge_first_section=merge_ranges)


@app.get("/", response_class=HTMLResponse)
async def index() -> HTMLResponse:
    html = """
    <!doctype html>
    <html lang="ru">
    <head>
        <meta charset="utf-8" />
        <title>Расчет бетона по остаткам</title>
        <meta name="viewport" content="width=device-width, initial-scale=1" />
        <style>
            * { box-sizing: border-box; }
            body {
                margin: 0;
                min-height: 100vh;
                font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                background: linear-gradient(180deg, #eef6ff 0%, #f8fbff 55%, #ffffff 100%);
                color: #10233f;
                display: flex;
                justify-content: flex-start;
                align-items: flex-start;
            }
            .wrap {
                width: 100%;
                max-width: 1680px;
                padding: 24px 24px 32px;
            }
            .page-layout {
                display: grid;
                grid-template-columns: 380px minmax(0, 1fr);
                gap: 20px;
                align-items: start;
            }
            .left-rail {
                position: sticky;
                top: 18px;
                align-self: start;
            }
            .left-stack {
                display: flex;
                flex-direction: column;
                gap: 18px;
            }
            .rail-section-title {
                font-size: 20px;
                font-weight: 700;
                margin: 0 0 10px;
                color: #123c73;
            }
            .help-btn-wrap {
                position: fixed;
                top: 20px;
                right: 24px;
                z-index: 10;
            }
            .help-btn {
                width: 28px;
                height: 28px;
                border-radius: 50%;
                border: 1px solid #a5c5ea;
                background: #eef6ff;
                color: #1d4ed8;
                font-size: 14px;
                font-weight: 600;
                cursor: pointer;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                padding: 0;
                line-height: 1;
            }
            .help-btn:hover {
                background: #dbeafe;
                color: #123c73;
            }
            .help-panel {
                position: fixed;
                inset: 0;
                background: rgba(18, 60, 115, 0.32);
                display: none;
                align-items: center;
                justify-content: center;
                z-index: 20;
            }
            .help-panel-inner {
                width: 100%;
                max-width: 840px;
                max-height: 90vh;
                background: linear-gradient(180deg, #ffffff 0%, #f7fbff 100%);
                border-radius: 16px;
                padding: 18px 20px;
                box-shadow: 0 20px 40px rgba(33, 93, 168, 0.18);
                display: flex;
                flex-direction: column;
                gap: 12px;
                border: 1px solid #d6e6fb;
                overflow: auto;
            }
            .help-panel h3 { margin: 0 0 6px; font-size: 14px; color: #123c73; }
            .help-panel p, .help-panel li { margin: 0 0 6px; font-size: 13px; line-height: 1.5; color: #334155; }
            .help-panel ul { margin: 0 0 10px; padding-left: 20px; }
            .help-shot-note {
                font-size: 12px;
                color: #64748b;
                margin: 0 0 10px;
                font-style: italic;
            }
            .help-shot-row {
                display: grid;
                grid-template-columns: 1fr;
                gap: 12px;
                margin: 8px 0 18px;
            }
            @media (min-width: 640px) {
                .help-shot-row { grid-template-columns: 1fr 1fr; }
            }
            .help-shot {
                border: 1px solid #cfe1f7;
                border-radius: 10px;
                padding: 10px 10px 12px;
                background: #f8fbff;
            }
            .help-shot-cap {
                font-size: 11px;
                font-weight: 600;
                color: #184a8b;
                margin: 0 0 8px;
            }
            .help-mock {
                border: 1px solid #bfdbfe;
                border-radius: 8px;
                background: #ffffff;
                padding: 8px;
                font-size: 10px;
                color: #64748b;
                line-height: 1.35;
            }
            .help-mock-title {
                font-weight: 600;
                color: #123c73;
                margin-bottom: 6px;
                font-size: 11px;
            }
            .help-mock-sub { font-size: 9px; color: #94a3b8; margin-bottom: 6px; }
            .help-mock-empty {
                min-height: 72px;
                display: flex;
                align-items: center;
                justify-content: center;
                text-align: center;
                border: 1px dashed #cbd5e1;
                border-radius: 6px;
                padding: 8px;
            }
            .help-mock-table {
                width: 100%;
                border-collapse: collapse;
                font-size: 9px;
            }
            .help-mock-table td {
                border-bottom: 1px solid #e2e8f0;
                padding: 3px 2px;
            }
            .help-mock-table td.num { text-align: right; color: #334155; }
            .help-mock-charts {
                display: grid;
                grid-template-columns: 1fr 1fr 1fr;
                gap: 4px;
                margin-top: 8px;
            }
            .help-mock-chart {
                height: 36px;
                background: linear-gradient(180deg, #93c5fd 55%, #e0f2fe 100%);
                border-radius: 4px;
                border: 1px solid #bfdbfe;
            }
            .help-mock-chart.help-mock-chart-alt {
                background: linear-gradient(180deg, #5eead4 50%, #ccfbf1 100%);
                border-color: #99f6e4;
            }
            .help-mock-chart.help-mock-chart-warn {
                background: linear-gradient(180deg, #fca5a5 40%, #fee2e2 100%);
                border-color: #fecaca;
            }
            .help-mock-grid {
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 2px;
                font-size: 8px;
                text-align: center;
            }
            .help-mock-cell {
                background: #f1f5f9;
                padding: 4px 2px;
                border-radius: 2px;
                min-height: 22px;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            .help-mock-cell.help-mock-hl {
                background: #dbeafe;
                font-weight: 600;
                color: #1e40af;
                border: 1px solid #93c5fd;
            }
            .help-mock-row2 {
                display: grid;
                grid-template-columns: 1fr 1.2fr;
                gap: 4px;
                font-size: 9px;
                margin-top: 4px;
            }
            .help-mock-label { color: #94a3b8; font-size: 8px; }
            .help-mock-input {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 4px 6px;
                color: #94a3b8;
            }
            .help-mock-input.help-mock-filled {
                background: #eff6ff;
                border-color: #93c5fd;
                color: #1e3a8a;
            }
            .help-mock-arrow {
                text-align: center;
                font-size: 18px;
                color: #94a3b8;
                margin: 4px 0;
            }
            @media (min-width: 640px) {
                .help-mock-arrow { display: none; }
            }
            .layout {
                display: flex;
                gap: 20px;
                align-items: flex-start;
                flex-wrap: wrap;
            }
            .stack-section {
                margin-top: 0;
            }
            .box {
                border-radius: 12px;
                border: 1px solid #cfe1f7;
                background: linear-gradient(180deg, #ffffff 0%, #f7fbff 100%);
                padding: 18px 18px 16px;
                position: relative;
                width: 100%;
                max-width: 380px;
                box-shadow: 0 14px 30px rgba(33, 93, 168, 0.08);
            }
            .result-box {
                flex: 1 1 0;
                min-width: 980px;
                border-radius: 12px;
                border: 1px solid #cfe1f7;
                background: linear-gradient(180deg, #ffffff 0%, #f7fbff 100%);
                padding: 14px 16px;
                font-size: 13px;
                box-shadow: 0 14px 30px rgba(33, 93, 168, 0.08);
                margin-top: 34px;
            }
            .result-box.has-result {
                margin-top: 0;
            }
            .result-title {
                font-size: 14px;
                font-weight: 600;
                margin-bottom: 6px;
            }
            .section-title {
                font-size: 22px;
                font-weight: 700;
                margin: 0 0 14px;
                color: #123c73;
            }
            .result-meta {
                font-size: 12px;
                color: #58708f;
                margin-bottom: 4px;
            }
            .result-ok {
                color: #1859a8;
            }
            .result-empty {
                color: #7f95b2;
            }
            .result-list {
                margin: 6px 0 0;
                padding-left: 16px;
            }
            .result-tables {
                margin-top: 14px;
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 14px;
                align-items: start;
            }
            .result-table-card {
                border-radius: 20px;
                border: 1px solid #cfe1f7;
                background: linear-gradient(180deg, #ffffff 0%, #f4f9ff 100%);
                box-shadow: 0 16px 34px rgba(33, 93, 168, 0.09);
                padding: 12px;
                min-width: 0;
            }
            .result-table-wrap {
                margin-top: 10px;
                border-radius: 14px;
                overflow: hidden;
                border: 1px solid #d8e8fb;
                background: #ffffff;
            }
            .result-table {
                width: 100%;
                border-collapse: collapse;
                font-size: 12px;
                table-layout: fixed;
            }
            .result-table th,
            .result-table td {
                padding: 9px 10px;
                border-bottom: 1px solid #e7f0fb;
            }
            .result-table th {
                background: linear-gradient(180deg, #e8f3ff 0%, #dcebff 100%);
                text-align: left;
                font-weight: 600;
                white-space: normal;
                color: #29558d;
                font-size: 11px;
                text-transform: uppercase;
                letter-spacing: 0.03em;
            }
            .result-table td {
                background: #ffffff;
                vertical-align: top;
                word-break: break-word;
            }
            .result-table th:nth-child(1),
            .result-table td:nth-child(1) { width: 40%; }
            .result-table th:nth-child(2),
            .result-table td:nth-child(2) { width: 15%; }
            .result-table th:nth-child(3),
            .result-table td:nth-child(3) { width: 22.5%; }
            .result-table th:nth-child(4),
            .result-table td:nth-child(4) { width: 22.5%; }
            .result-table tr:nth-child(even) td {
                background: #f7fbff;
            }
            .result-table tr:last-child td {
                border-bottom: none;
            }
            .result-table tr.row-max td {
                background: #e7f1ff !important;
                font-weight: 600;
            }
            .result-section-title {
                margin-top: 0;
                margin-bottom: 6px;
                font-size: 14px;
                font-weight: 700;
                color: #123c73;
            }
            .result-section-subtitle {
                color: #6580a3;
                font-size: 12px;
                margin-bottom: 2px;
            }
            .result-badge {
                display: inline-flex;
                align-items: center;
                margin-left: 8px;
                padding: 2px 8px;
                border-radius: 999px;
                background: #dbeafe;
                color: #1d4ed8;
                font-size: 11px;
                font-weight: 700;
                white-space: nowrap;
            }
            .result-name {
                font-weight: 600;
                color: #10233f;
            }
            .result-limiter {
                margin-top: 4px;
                font-size: 11px;
                line-height: 1.35;
                color: #5f7593;
            }
            .result-num {
                text-align: center;
                vertical-align: middle;
                white-space: nowrap;
                font-variant-numeric: tabular-nums;
            }
            .future-box {
                margin-top: 20px;
                border-radius: 18px;
                border: 1px solid #e5e7eb;
                background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
                padding: 18px 20px;
                box-shadow: 0 10px 24px rgba(15, 23, 42, 0.05);
            }
            .future-title {
                font-size: 16px;
                font-weight: 700;
                margin-bottom: 6px;
                color: #0f172a;
            }
            .future-text {
                font-size: 13px;
                color: #64748b;
                max-width: 720px;
            }
            @media (max-width: 980px) {
                .page-layout {
                    grid-template-columns: 1fr;
                }
                .left-rail {
                    position: static;
                }
                .result-box {
                    min-width: 0;
                    margin-top: 0;
                }
                .result-tables {
                    grid-template-columns: 1fr;
                }
            }
            h1 {
                font-size: 18px;
                margin: 0 0 12px;
                font-weight: 600;
                color: #123c73;
            }
            .field {
                margin-bottom: 12px;
            }
            label {
                display: block;
                font-size: 12px;
                margin-bottom: 6px;
                color: #5a7696;
            }
            input[type="file"] {
                width: 100%;
                border-radius: 8px;
                border: 1px solid #c7dcf5;
                padding: 8px 10px;
                font-size: 14px;
                background: #ffffff;
            }
            input[type="file"]::file-selector-button {
                border: 1px solid #bfd9f8;
                border-radius: 6px;
                padding: 6px 10px;
                margin-right: 8px;
                background: #eaf4ff;
                color: #1d4ed8;
                font-size: 12px;
                cursor: pointer;
            }
            .btn {
                margin-top: 4px;
                width: 100%;
                border-radius: 8px;
                border: 1px solid #bfd9f8;
                padding: 8px 10px;
                font-weight: 500;
                font-size: 13px;
                color: #184a8b;
                background: #eaf4ff;
                cursor: pointer;
            }
            .btn-row {
                display: flex;
                gap: 8px;
                margin-top: 4px;
            }
            .btn-row .btn {
                margin-top: 0;
            }
            .hp-field {
                position: absolute;
                left: -9999px;
                opacity: 0;
                pointer-events: none;
            }
            .field select {
                width: 100%;
                border-radius: 8px;
                border: 1px solid #c7dcf5;
                padding: 8px 10px;
                font-size: 14px;
                background: #ffffff;
                color: #10233f;
            }
            .cfg-btn {
                position: absolute;
                top: 12px;
                right: 12px;
                width: 44px;
                height: 44px;
                border-radius: 12px;
                border: 1px solid #c7dcf5;
                background: #eff7ff;
                cursor: pointer;
                display: flex;
                align-items: center;
                justify-content: center;
                color: #1d4ed8;
                font-size: 22px;
                line-height: 1;
                z-index: 10;
            }
            .cfg-btn:hover {
                color: #123c73;
                background: #e3f0ff;
            }
            .cfg-panel {
                position: fixed;
                inset: 0;
                background: rgba(18, 60, 115, 0.32);
                display: none;
                align-items: center;
                justify-content: center;
                z-index: 20;
            }
            .cfg-panel-inner {
                width: 100%;
                max-width: 720px;
                max-height: 90vh;
                background: linear-gradient(180deg, #ffffff 0%, #f7fbff 100%);
                border-radius: 16px;
                padding: 16px 18px;
                box-shadow: 0 20px 40px rgba(33, 93, 168, 0.18);
                display: flex;
                flex-direction: column;
                gap: 12px;
                border: 1px solid #d6e6fb;
            }
            .cfg-header {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 8px;
            }
            .cfg-title {
                font-size: 15px;
                font-weight: 600;
            }
            .cfg-close {
                border: none;
                background: transparent;
                cursor: pointer;
                font-size: 18px;
                line-height: 1;
                color: #6b7280;
            }
            .cfg-body {
                font-size: 12px;
                display: flex;
                flex-direction: column;
                gap: 8px;
            }
            .cfg-tabs {
                display: flex;
                gap: 2px;
                border-bottom: 1px solid #e5e7eb;
            }
            .cfg-tab {
                padding: 6px 12px;
                border: none;
                background: transparent;
                cursor: pointer;
                font-size: 12px;
                color: #6b7fa0;
                border-bottom: 2px solid transparent;
                margin-bottom: -1px;
            }
            .cfg-tab:hover { color: #123c73; }
            .cfg-tab.active {
                color: #1d4ed8;
                border-bottom-color: #2563eb;
                font-weight: 500;
            }
            .cfg-pane { display: none; overflow: auto; max-height: 50vh; }
            .cfg-pane.active { display: block; }
            .cfg-section label {
                display: block;
                margin-bottom: 4px;
                font-size: 11px;
                color: #5a7696;
            }
            .cfg-table-wrap { overflow-x: auto; }
            .cfg-table {
                width: 100%;
                border-collapse: collapse;
                font-size: 12px;
            }
            .cfg-table th, .cfg-table td {
                border: 1px solid #dbe8f7;
                padding: 4px 6px;
                text-align: left;
            }
            .cfg-table th { background: #eef6ff; font-weight: 500; color: #29558d; }
            .cfg-table input, .cfg-table select {
                width: 100%;
                border: 1px solid #c7dcf5;
                border-radius: 4px;
                padding: 4px 6px;
                font-size: 12px;
                color: #10233f;
                background: #ffffff;
            }
            .cfg-table .col-del { width: 28px; text-align: center; }
            .cfg-recipe-block {
                border: 1px solid #d6e6fb;
                border-radius: 8px;
                padding: 8px;
                margin-bottom: 8px;
                background: linear-gradient(180deg, #ffffff 0%, #f4f9ff 100%);
            }
            .cfg-recipe-block h4 { margin: 0 0 6px; font-size: 12px; }
            .cfg-recipe-name { margin-bottom: 6px; }
            .cfg-recipe-name input { width: 100%; max-width: 280px; padding: 4px 6px; font-size: 12px; border-radius: 4px; border: 1px solid #c7dcf5; color: #10233f; background: #ffffff; }
            .cfg-add-row { margin-top: 6px; }
            .cfg-btn-sm {
                border: none;
                background: transparent;
                cursor: pointer;
                padding: 2px 6px;
                font-size: 11px;
                color: #6480a1;
            }
            .cfg-btn-sm:hover { color: #dc2626; }
            .cfg-footer {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 8px;
                margin-top: 6px;
            }
            .cfg-footer-left,
            .cfg-footer-right {
                display: flex;
                align-items: center;
                gap: 6px;
                flex-wrap: wrap;
            }
            .cfg-input {
                border-radius: 6px;
                border: 1px solid #c7dcf5;
                padding: 4px 6px;
                font-size: 12px;
                color: #10233f;
                background: #ffffff;
            }
            .cfg-select {
                border-radius: 6px;
                border: 1px solid #c7dcf5;
                padding: 4px 6px;
                font-size: 12px;
                color: #10233f;
                background: #ffffff;
            }
            .cfg-btn-prim,
            .cfg-btn-sec,
            .cfg-btn-danger {
                border-radius: 6px;
                border: none;
                padding: 5px 8px;
                font-size: 11px;
                cursor: pointer;
            }
            .cfg-btn-prim {
                background: #eaf4ff;
                color: #184a8b;
                border: 1px solid #bfd9f8;
            }
            .cfg-btn-sec {
                background: #f3f8fe;
                color: #123c73;
                border: 1px solid #d7e6f8;
            }
            .cfg-btn-danger {
                background: #fff1f2;
                color: #be123c;
                border: 1px solid #fecdd3;
            }
            .cfg-hint {
                font-size: 11px;
                color: #5a7696;
            }
            .cfg-status {
                display: none;
                margin-top: 8px;
                padding: 8px 10px;
                border-radius: 8px;
                border: 1px solid #fecdd3;
                background: #fff1f2;
                color: #9f1239;
                font-size: 12px;
                line-height: 1.45;
                white-space: pre-line;
            }
            .cfg-table tr.cfg-error-row td {
                background: #fff7f7;
            }
            .cfg-table tr.cfg-error-row input,
            .cfg-table tr.cfg-error-row select,
            .cfg-recipe-block.cfg-error-block input,
            .cfg-recipe-block.cfg-error-block select,
            .cfg-error-field,
            .cfg-input.cfg-error-input {
                border-color: #f43f5e;
                background: #fff7f7;
            }
            .cfg-recipe-block.cfg-error-block {
                border-color: #f43f5e;
                box-shadow: inset 0 0 0 1px #fecdd3;
            }
            .chart-preview {
                margin-top: 22px;
                padding-top: 18px;
                border-top: 1px solid #c7dcf5;
            }
            .chart-preview-head {
                margin-bottom: 14px;
            }
            .chart-preview-title {
                font-size: 15px;
                font-weight: 600;
                color: #10233f;
                margin-bottom: 4px;
            }
            .chart-preview-hint {
                font-size: 12px;
                color: #64748b;
                line-height: 1.45;
            }
            .chart-preview-grid {
                display: grid;
                grid-template-columns: 1fr;
                gap: 14px;
            }
            @media (min-width: 900px) {
                .chart-preview-grid {
                    grid-template-columns: 1fr 1fr;
                }
            }
            .chart-preview-card {
                background: linear-gradient(165deg, #f8fbff 0%, #ffffff 55%);
                border: 1px solid #d7e6f8;
                border-radius: 12px;
                padding: 12px 12px 8px;
                box-shadow: 0 8px 24px rgba(33, 93, 168, 0.06);
            }
            .chart-preview-cap {
                font-size: 12px;
                font-weight: 600;
                color: #184a8b;
                margin-bottom: 8px;
            }
            .chart-preview-canvas {
                position: relative;
                height: 240px;
            }
        </style>
    </head>
    <body>
        <div class="wrap">
            <div class="help-btn-wrap"><button type="button" class="help-btn" id="helpBtn" title="Справка">?</button></div>
            <div class="page-layout">
                <div class="left-rail">
                    <div class="left-stack">
""" + _build_left_stack_html() + """
                    </div>
                </div>
                <div class="result-box result-empty" id="resultBox">
                    <div class="result-title">Результат расчета</div>
                    <div>Справа показывается результат того раздела, который вы посчитали последним: бетон или ЖБИ.</div>
                </div>
            </div>
        </div>
        <div class="cfg-panel" id="cfgPanel">
            <div class="cfg-panel-inner">
                <div class="cfg-header">
                    <div>
                        <div class="cfg-title" id="cfgTitle">Настройки конфигурации</div>
                        <div class="cfg-hint">Редактирование работает только для веб-сервиса, бот использует базовую конфигурацию.</div>
                    </div>
                    <button type="button" class="cfg-close" id="cfgClose" aria-label="Закрыть">×</button>
                </div>
                <div class="cfg-footer">
                    <div class="cfg-footer-left">
                        <select id="cfgProfileSelect" class="cfg-select">
                            <option value="__base__">Базовая конфигурация</option>
                        </select>
                        <button type="button" class="cfg-btn-sec" id="cfgProfileApply">Выбрать</button>
                        <button type="button" class="cfg-btn-danger" id="cfgProfileDelete">Удалить</button>
                    </div>
                    <div class="cfg-footer-right">
                        <input id="cfgProfileName" class="cfg-input" placeholder="Имя профиля" />
                        <button type="button" class="cfg-btn-prim" id="cfgProfileSave">Сохранить профиль</button>
                    </div>
                </div>
                <div class="cfg-status" id="cfgStatus"></div>
                <div class="cfg-body">
                    <div class="cfg-tabs">
                        <button type="button" class="cfg-tab active" data-tab="materials">Материалы</button>
                        <button type="button" class="cfg-tab" data-tab="recipes">Составы</button>
                        <button type="button" class="cfg-tab" data-tab="prices">Цены</button>
                    </div>
                    <div id="cfgPaneMaterials" class="cfg-pane active">
                        <div class="cfg-section">
                            <label>Наименование и варианты написания</label>
                            <div class="cfg-table-wrap">
                                <table class="cfg-table" id="cfgMaterialsTable">
                                    <thead><tr><th>Наименование</th><th>Варианты написания (через ||)</th><th class="col-del"></th></tr></thead>
                                    <tbody id="cfgMaterialsBody"></tbody>
                                </table>
                            </div>
                            <button type="button" class="cfg-btn-sec cfg-add-row" id="cfgMaterialsAdd">+ Добавить материал</button>
                        </div>
                    </div>
                    <div id="cfgPaneRecipes" class="cfg-pane">
                        <div class="cfg-section">
                            <label>Позиции и их составляющие (кг на 1 м³)</label>
                            <div id="cfgRecipesList"></div>
                            <button type="button" class="cfg-btn-sec cfg-add-row" id="cfgRecipesAdd">+ Добавить состав</button>
                        </div>
                    </div>
                    <div id="cfgPanePrices" class="cfg-pane">
                        <div class="cfg-section">
                            <label>Цены по наименованиям</label>
                            <div class="cfg-table-wrap">
                                <table class="cfg-table" id="cfgPricesTable">
                                    <thead><tr><th>Наименование</th><th>Без доставки без НДС</th><th>Без доставки с НДС 22%</th><th>Самовывоз без НДС</th><th>Самовывоз с НДС 22%</th><th class="col-del"></th></tr></thead>
                                    <tbody id="cfgPricesBody"></tbody>
                                </table>
                            </div>
                            <button type="button" class="cfg-btn-sec cfg-add-row" id="cfgPricesAdd">+ Добавить цену</button>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="help-panel" id="helpPanel">
            <div class="help-panel-inner">
                <div class="cfg-header">
                    <div class="cfg-title">Справка</div>
                    <button type="button" class="cfg-close" id="helpClose" aria-label="Закрыть">×</button>
                </div>
                <div class="help-body">
                    <p class="help-shot-note">Ниже — условные схемы экрана (не фотографии), чтобы было видно «где что» до и после действий.</p>

                    <h3>1. Область результата справа</h3>
                    <p>Слева — форма и файл; справа — результат. Пока не нажали «Посчитать», блок пустой или с подсказкой. После расчёта — таблицы и блок аналитики с графиками.</p>
                    <div class="help-shot-row">
                        <div class="help-shot">
                            <div class="help-shot-cap">До: расчёт не запускали</div>
                            <div class="help-mock">
                                <div class="help-mock-title">Результат расчета</div>
                                <div class="help-mock-sub">правая колонка страницы</div>
                                <div class="help-mock-empty">Справа показывается результат того раздела, который вы посчитали последним…</div>
                            </div>
                        </div>
                        <div class="help-shot">
                            <div class="help-shot-cap">После: «Посчитать»</div>
                            <div class="help-mock">
                                <div class="help-mock-title">Результат расчета</div>
                                <table class="help-mock-table">
                                    <tr><td>Бетон В25</td><td class="num">12,5</td><td class="num">… ₽</td></tr>
                                    <tr><td>Бетон В30</td><td class="num">8,0</td><td class="num">… ₽</td></tr>
                                </table>
                                <div class="help-mock-sub" style="margin-top:6px;">Аналитика</div>
                                <div class="help-mock-charts">
                                    <div class="help-mock-chart" title="объёмы"></div>
                                    <div class="help-mock-chart help-mock-chart-alt" title="цены"></div>
                                    <div class="help-mock-chart help-mock-chart-warn" title="запасы"></div>
                                </div>
                            </div>
                        </div>
                    </div>

                    <h3>2. Файл Excel: сальдо</h3>
                    <p>В ведомости нужен столбец с сальдо на конец периода — по нему берутся килограммы для расчёта.</p>
                    <div class="help-shot-row">
                        <div class="help-shot">
                            <div class="help-shot-cap">До: неочевидно, какой столбец важен</div>
                            <div class="help-mock">
                                <div class="help-mock-grid">
                                    <div class="help-mock-cell">Наимен.</div>
                                    <div class="help-mock-cell">Дебет</div>
                                    <div class="help-mock-cell">Кредит</div>
                                    <div class="help-mock-cell">Сальдо</div>
                                    <div class="help-mock-cell">Цемент</div>
                                    <div class="help-mock-cell">…</div>
                                    <div class="help-mock-cell">…</div>
                                    <div class="help-mock-cell">1200</div>
                                </div>
                            </div>
                        </div>
                        <div class="help-shot">
                            <div class="help-shot-cap">После: ищем колонку «Сальдо на конец периода»</div>
                            <div class="help-mock">
                                <div class="help-mock-grid">
                                    <div class="help-mock-cell">Наимен.</div>
                                    <div class="help-mock-cell">…</div>
                                    <div class="help-mock-cell help-mock-hl" style="grid-column:span 2;">Сальдо на конец периода</div>
                                    <div class="help-mock-cell">Цемент</div>
                                    <div class="help-mock-cell">…</div>
                                    <div class="help-mock-cell">…</div>
                                    <div class="help-mock-cell help-mock-hl">1200</div>
                                </div>
                                <div class="help-mock-sub" style="margin-top:6px;">Эти числа (кг) уходят в калькулятор.</div>
                            </div>
                        </div>
                    </div>

                    <h3>3. Настройки: материал и варианты написания</h3>
                    <p>Шестерёнка → пароль → вкладка «Материалы». В одной ячейке можно несколько вариантов через <strong>||</strong>, как в Excel встречается название.</p>
                    <div class="help-shot-row">
                        <div class="help-shot">
                            <div class="help-shot-cap">До: только одно имя</div>
                            <div class="help-mock">
                                <div class="help-mock-row2">
                                    <div>
                                        <div class="help-mock-label">Материал</div>
                                        <div class="help-mock-input">Цемент</div>
                                    </div>
                                    <div>
                                        <div class="help-mock-label">Варианты (||)</div>
                                        <div class="help-mock-input">— пусто —</div>
                                    </div>
                                </div>
                            </div>
                        </div>
                        <div class="help-shot">
                            <div class="help-shot-cap">После: все строки из файла, которые считаем этим материалом</div>
                            <div class="help-mock">
                                <div class="help-mock-row2">
                                    <div>
                                        <div class="help-mock-label">Материал</div>
                                        <div class="help-mock-input help-mock-filled">Цемент</div>
                                    </div>
                                    <div>
                                        <div class="help-mock-label">Варианты (||)</div>
                                        <div class="help-mock-input help-mock-filled">Цемент||Портландцемент</div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>

                    <h3>Расчёт бетона</h3>
                    <p>Для каждого типа бетона — <strong>максимум м³</strong> по остаткам: для каждого материала лимит = остаток ÷ расход на 1 м³; по рецепту берётся минимум. Стоимость = максимум м³ × цена за 1 м³ (настройки).</p>

                    <h3>Расчёт ЖБИ</h3>
                    <p>Сначала доступный бетон, затем по каждому изделию — максимум штук; ограничивает самый «узкий» ресурс. Стоимость = шт × цена за шт. «Скачать» — Excel с таблицей и ценами в четырёх вариантах, как для бетона.</p>
                </div>
            </div>
        </div>
        <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js" crossorigin="anonymous"></script>
        <script>
        document.addEventListener('DOMContentLoaded', function() {
            var cfgBtn = document.getElementById('cfgBtn');
            var jbiCfgBtn = document.getElementById('jbiCfgBtn');
            var cfgPanel = document.getElementById('cfgPanel');

            var cfgClose = document.getElementById('cfgClose');
            var cfgTitle = document.getElementById('cfgTitle');
            var sel = document.getElementById('cfgProfileSelect');
            var saveBtn = document.getElementById('cfgProfileSave');
            var applyBtn = document.getElementById('cfgProfileApply');
            var delBtn = document.getElementById('cfgProfileDelete');
            var nameInput = document.getElementById('cfgProfileName');
            var cfgStatus = document.getElementById('cfgStatus');
            var mainProfileSelect = document.getElementById('mainProfileSelect');
            var cfgMaterialsBody = document.getElementById('cfgMaterialsBody');
            var cfgRecipesList = document.getElementById('cfgRecipesList');
            var cfgPricesBody = document.getElementById('cfgPricesBody');
            var calcForm = document.getElementById('calcForm');
            var resultBox = document.getElementById('resultBox');
            var calcOnlyBtn = document.getElementById('calcOnlyBtn');
            var downloadBtn = document.getElementById('downloadBtn');
            var jbiForm = document.getElementById('jbiForm');
            var jbiProfileSelect = document.getElementById('jbiProfileSelect');
            var jbiCalcOnlyBtn = document.getElementById('jbiCalcOnlyBtn');
            var jbiDownloadBtn = document.getElementById('jbiDownloadBtn');

            if (!cfgBtn || !cfgPanel) {
                return;
            }

            var currentMaterialNames = [];
            var currentExternalMaterialNames = [];
            var currentConfigScope = 'beton';
            var currentPanelPassword = null;
            var __previewChartInstances = [];

            function clearConfigStatus() {
                if (!cfgStatus) return;
                cfgStatus.style.display = 'none';
                cfgStatus.textContent = '';
            }
            function showConfigStatus(message) {
                if (!message) {
                    clearConfigStatus();
                    return;
                }
                if (!cfgStatus) {
                    alert(message);
                    return;
                }
                cfgStatus.textContent = message;
                cfgStatus.style.display = 'block';
            }
            function clearConfigErrors() {
                clearConfigStatus();
                if (nameInput) nameInput.classList.remove('cfg-error-input');
                var fields = cfgPanel.querySelectorAll('.cfg-error-field');
                for (var f = 0; f < fields.length; f++) fields[f].classList.remove('cfg-error-field');
                if (cfgMaterialsBody) {
                    var matRows = cfgMaterialsBody.querySelectorAll('tr');
                    for (var i = 0; i < matRows.length; i++) matRows[i].classList.remove('cfg-error-row');
                }
                if (cfgPricesBody) {
                    var priceRows = cfgPricesBody.querySelectorAll('tr');
                    for (var j = 0; j < priceRows.length; j++) priceRows[j].classList.remove('cfg-error-row');
                }
                if (cfgRecipesList) {
                    var recipeBlocks = cfgRecipesList.querySelectorAll('.cfg-recipe-block');
                    for (var k = 0; k < recipeBlocks.length; k++) recipeBlocks[k].classList.remove('cfg-error-block');
                }
            }
            function reindexConfigUI() {
                if (cfgMaterialsBody) {
                    var matRows = cfgMaterialsBody.querySelectorAll('tr');
                    for (var i = 0; i < matRows.length; i++) matRows[i].setAttribute('data-idx', String(i));
                }
                if (cfgPricesBody) {
                    var priceRows = cfgPricesBody.querySelectorAll('tr');
                    for (var j = 0; j < priceRows.length; j++) priceRows[j].setAttribute('data-idx', String(j));
                }
                if (cfgRecipesList) {
                    var recipeBlocks = cfgRecipesList.querySelectorAll('.cfg-recipe-block');
                    for (var k = 0; k < recipeBlocks.length; k++) recipeBlocks[k].setAttribute('data-idx', String(k));
                }
            }
            function humanizeValidationMessage(error) {
                var msg = error && error.message ? String(error.message) : 'Проверьте выделенные поля.';
                msg = msg.replace('нужен хотя бы один алиас', 'нужно указать хотя бы один вариант написания');
                return msg;
            }
            function applyValidationErrors(errors) {
                clearConfigErrors();
                reindexConfigUI();
                if (!errors || !errors.length) return;
                var firstEl = null;
                var firstTab = null;
                var messages = [];
                function markField(el) {
                    if (el) el.classList.add('cfg-error-field');
                }
                function pickFirst(container, selector) {
                    return container && container.querySelector ? container.querySelector(selector) : null;
                }
                for (var i = 0; i < errors.length; i++) {
                    var error = errors[i] || {};
                    var msg = error && error.message ? String(error.message) : '';
                    var idx = typeof error.index === 'number' ? error.index : -1;
                    messages.push('• ' + humanizeValidationMessage(error));
                    if (error.field === 'name') {
                        if (!firstTab) firstTab = 'materials';
                        if (nameInput) {
                            nameInput.classList.add('cfg-error-input');
                            markField(nameInput);
                            if (!firstEl) firstEl = nameInput;
                        }
                        continue;
                    }
                    if (error.field === 'materials') {
                        if (!firstTab) firstTab = 'materials';
                        if (cfgMaterialsBody && idx >= 0) {
                            var matRow = cfgMaterialsBody.querySelector('tr[data-idx="' + idx + '"]');
                            if (matRow) {
                                matRow.classList.add('cfg-error-row');
                                var materialField = null;
                                if (msg.indexOf('алис') >= 0 || msg.indexOf('алиас') >= 0 || msg.indexOf('вариант') >= 0 || msg.indexOf('конфликтует') >= 0) {
                                    materialField = pickFirst(matRow, '.mat-aliases');
                                } else {
                                    materialField = pickFirst(matRow, '.mat-name');
                                }
                                markField(materialField);
                                if (!firstEl) firstEl = materialField || pickFirst(matRow, 'input');
                            }
                        }
                        continue;
                    }
                    if (error.field === 'recipes') {
                        if (!firstTab) firstTab = 'recipes';
                        if (cfgRecipesList && idx >= 0) {
                            var block = cfgRecipesList.querySelector('.cfg-recipe-block[data-idx="' + idx + '"]');
                            if (block) {
                                block.classList.add('cfg-error-block');
                                var recipeField = null;
                                if (msg.indexOf('Дублируется состав') >= 0 || msg.indexOf('должно быть наименование') >= 0) {
                                    recipeField = pickFirst(block, '.rec-name');
                                    markField(recipeField);
                                } else if (msg.indexOf('неизвестный материал') >= 0 || msg.indexOf('пустой материал') >= 0) {
                                    var materialSelects = block.querySelectorAll('.rec-mat-name');
                                    for (var rs = 0; rs < materialSelects.length; rs++) markField(materialSelects[rs]);
                                    recipeField = materialSelects.length ? materialSelects[0] : null;
                                } else if (msg.indexOf('нечисловое значение') >= 0 || msg.indexOf('отрицательное значение') >= 0 || msg.indexOf('расходом больше нуля') >= 0) {
                                    var kgFields = block.querySelectorAll('.rec-mat-kg');
                                    for (var kgIdx = 0; kgIdx < kgFields.length; kgIdx++) markField(kgFields[kgIdx]);
                                    recipeField = kgFields.length ? kgFields[0] : null;
                                } else {
                                    recipeField = pickFirst(block, '.rec-name');
                                    markField(recipeField);
                                }
                                if (!firstEl) firstEl = recipeField || pickFirst(block, 'input, select');
                            }
                        }
                        continue;
                    }
                    if (error.field === 'prices') {
                        if (!firstTab) firstTab = 'prices';
                        if (cfgPricesBody && idx >= 0) {
                            var priceRow = cfgPricesBody.querySelector('tr[data-idx="' + idx + '"]');
                            if (priceRow) {
                                priceRow.classList.add('cfg-error-row');
                                var priceField = null;
                                if (msg.indexOf('no_delivery_no_vat') >= 0) priceField = pickFirst(priceRow, '.price-nd-nv');
                                else if (msg.indexOf('no_delivery_vat_22') >= 0) priceField = pickFirst(priceRow, '.price-nd-v');
                                else if (msg.indexOf('pickup_no_vat') >= 0) priceField = pickFirst(priceRow, '.price-pick-nv');
                                else if (msg.indexOf('pickup_vat_22') >= 0) priceField = pickFirst(priceRow, '.price-pick-v');
                                else if (msg.indexOf('У цены должно быть наименование') >= 0 || msg.indexOf('Дублируется цена') >= 0) priceField = pickFirst(priceRow, '.price-name');
                                else priceField = pickFirst(priceRow, 'input');
                                markField(priceField);
                                if (!firstEl) firstEl = priceField || pickFirst(priceRow, 'input');
                            }
                        }
                    }
                }
                if (firstTab) setActiveTab(firstTab);
                showConfigStatus('Проверьте выделенные строки и исправьте их:\\n' + messages.join('\\n'));
                if (firstEl && firstEl.focus) firstEl.focus();
            }
            function parseResponseError(res, fallbackMessage) {
                return res.text().then(function(text) {
                    var payload = null;
                    try {
                        payload = text ? JSON.parse(text) : null;
                    } catch (e) {}
                    var detail = payload && payload.detail != null ? payload.detail : payload;
                    var message = fallbackMessage || 'Произошла ошибка.';
                    var errors = [];
                    if (detail && typeof detail === 'object') {
                        if (detail.message) message = detail.message;
                        if (detail.errors && detail.errors.length) errors = detail.errors;
                    } else if (typeof detail === 'string' && detail) {
                        message = detail;
                    } else if (text) {
                        message = text;
                    }
                    return { message: message, errors: errors };
                });
            }

            function setActiveTab(tabKey) {
                var tabs = document.querySelectorAll('.cfg-tab');
                var panes = document.querySelectorAll('.cfg-pane');
                for (var i = 0; i < tabs.length; i++) tabs[i].classList.remove('active');
                for (var j = 0; j < panes.length; j++) panes[j].classList.remove('active');
                for (var k = 0; k < tabs.length; k++) {
                    if (tabs[k].getAttribute('data-tab') === tabKey) tabs[k].classList.add('active');
                }
                var paneId = 'cfgPane' + (tabKey ? tabKey.charAt(0).toUpperCase() + tabKey.slice(1) : '');
                var pane = document.getElementById(paneId);
                if (pane) pane.classList.add('active');
            }

            var tabButtons = document.querySelectorAll('.cfg-tab');
            for (var t = 0; t < tabButtons.length; t++) {
                tabButtons[t].addEventListener('click', function() {
                    setActiveTab(this.getAttribute('data-tab'));
                });
            }

            function getScopeTitle(scope) {
                return scope === 'jbi' ? 'Настройки конфигурации ЖБИ' : 'Настройки конфигурации бетона';
            }
            function getProfileLabel(scope) {
                return scope === 'jbi' ? 'Изделия и их составляющие (кг на 1 м³)' : 'Виды бетона и их составляющие (кг на 1 м³)';
            }
            function askConfigPassword() {
                var entered = window.prompt('Введите пароль конфигуратора');
                if (!entered) return null;
                return String(entered);
            }
            function configFetch(url, options, passwordOverride) {
                options = options || {};
                var password = passwordOverride != null ? passwordOverride : currentPanelPassword;
                if (!password) {
                    password = askConfigPassword();
                }
                if (!password) {
                    return Promise.reject(new Error('Пароль не введен'));
                }
                if (/[^\\u0000-\\u00FF]/.test(password)) {
                    return Promise.reject(new Error('Пароль можно вводить только цифрами, латиницей и обычными символами.'));
                }
                options.headers = options.headers || {};
                options.headers['X-Config-Password'] = password;
                return fetch(url, options).then(function(res) {
                    if (res.status === 401) {
                        currentPanelPassword = null;
                    }
                    return res;
                });
            }
            function openPanel(scope) {
                currentConfigScope = scope === 'jbi' ? 'jbi' : 'beton';
                if (cfgTitle) cfgTitle.textContent = getScopeTitle(currentConfigScope);
                var password = askConfigPassword();
                if (!password) return;
                currentPanelPassword = password;
                loadConfig(password);
            }
            function closePanel() {
                currentPanelPassword = null;
                clearConfigErrors();
                cfgPanel.style.display = 'none';
            }

            function fillProfileSelect(selectEl, profiles, active) {
                if (!selectEl) return;
                selectEl.innerHTML = '<option value="__base__">По умолчанию</option>';
                for (var i = 0; i < profiles.length; i++) {
                    var opt = document.createElement('option');
                    opt.value = profiles[i].name;
                    opt.textContent = profiles[i].name;
                    selectEl.appendChild(opt);
                }
                selectEl.value = active || '__base__';
            }
            function loadProfileOptions(scope, selectEl) {
                if (!selectEl || !window.fetch) return;
                return fetch('/api/config/options?scope=' + encodeURIComponent(scope))
                    .then(function(res) {
                        if (!res.ok) return null;
                        return res.json();
                    })
                    .then(function(data) {
                        if (!data) return;
                        fillProfileSelect(selectEl, data.profiles || [], data.active_profile || '__base__');
                    })
                    .catch(function(e) {
                        console.error(e);
                    });
            }
            function loadMainProfileSelects() {
                loadProfileOptions('beton', mainProfileSelect);
                loadProfileOptions('jbi', jbiProfileSelect);
            }

            loadMainProfileSelects();

            cfgBtn.addEventListener('click', function() { openPanel('beton'); });
            if (jbiCfgBtn) jbiCfgBtn.addEventListener('click', function() { openPanel('jbi'); });

            if (cfgClose) cfgClose.addEventListener('click', closePanel);
            cfgPanel.addEventListener('click', function(e) {
                if (e.target === cfgPanel) closePanel();
            });

            var helpBtn = document.getElementById('helpBtn');
            var helpPanel = document.getElementById('helpPanel');
            var helpClose = document.getElementById('helpClose');
            if (helpBtn && helpPanel) {
                helpBtn.addEventListener('click', function() { helpPanel.style.display = 'flex'; });
                if (helpClose) helpClose.addEventListener('click', function() { helpPanel.style.display = 'none'; });
                helpPanel.addEventListener('click', function(e) {
                    if (e.target === helpPanel) helpPanel.style.display = 'none';
                });
            }
            cfgPanel.addEventListener('input', function(e) {
                var row = safeClosest(e.target, 'tr');
                var block = safeClosest(e.target, '.cfg-recipe-block');
                if (row) row.classList.remove('cfg-error-row');
                if (block) block.classList.remove('cfg-error-block');
                if (e.target === nameInput) nameInput.classList.remove('cfg-error-input');
            });

            function escapeAttr(s) {
                return String(s || '').replace(/"/g, '&quot;');
            }
            function escapeHtml(s) {
                return String(s || '').replace(/</g, '&lt;');
            }
            function formatNumber(value, digits) {
                if (value == null || isNaN(value)) return '—';
                var fixed = Number(value).toFixed(digits);
                var parts = fixed.split('.');
                parts[0] = parts[0].replace(/\B(?=(\d{3})+(?!\d))/g, ' ');
                return parts.join('.');
            }
            function setScopeTexts() {
                var lbl = document.querySelector('#cfgPaneRecipes .cfg-section label');
                var btn = document.getElementById('cfgRecipesAdd');
                if (lbl) lbl.textContent = getProfileLabel(currentConfigScope);
                if (btn) btn.textContent = currentConfigScope === 'jbi' ? '+ Добавить состав ЖБИ' : '+ Добавить состав';
            }
            function getAvailableMaterialNames() {
                var all = [];
                var seen = {};
                for (var i = 0; i < currentMaterialNames.length; i++) {
                    if (!seen[currentMaterialNames[i]]) {
                        seen[currentMaterialNames[i]] = true;
                        all.push(currentMaterialNames[i]);
                    }
                }
                for (var j = 0; j < currentExternalMaterialNames.length; j++) {
                    if (!seen[currentExternalMaterialNames[j]]) {
                        seen[currentExternalMaterialNames[j]] = true;
                        all.push(currentExternalMaterialNames[j]);
                    }
                }
                return all;
            }

            function renderMaterials(arr) {
                arr = arr || [];
                currentMaterialNames = [];
                for (var i = 0; i < arr.length; i++) currentMaterialNames.push(arr[i].name || '');
                if (!cfgMaterialsBody) return;
                cfgMaterialsBody.innerHTML = '';
                for (var j = 0; j < arr.length; j++) {
                    var m = arr[j] || {};
                    var tr = document.createElement('tr');
                    tr.setAttribute('data-idx', String(j));
                    var aliases = Array.isArray(m.aliases) ? m.aliases.join(' || ') : '';
                    tr.innerHTML =
                        '<td><input type="text" class="mat-name" value="' +
                        escapeAttr(m.name || '') +
                        '" /></td><td><input type="text" class="mat-aliases" value="' +
                        escapeAttr(aliases) +
                        '" placeholder="через ||" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-mat" title="Удалить">✕</button></td>';
                    cfgMaterialsBody.appendChild(tr);
                }
            }

            function renderRecipes(arr, materialNames) {
                arr = arr || [];
                var names = materialNames || getAvailableMaterialNames();
                if (!cfgRecipesList) return;
                cfgRecipesList.innerHTML = '';
                for (var i = 0; i < arr.length; i++) {
                    var rec = arr[i] || {};
                    var block = document.createElement('div');
                    block.className = 'cfg-recipe-block';
                    block.setAttribute('data-idx', String(i));
                    var matsHtml = '';
                    var mats = rec.materials && typeof rec.materials === 'object' ? rec.materials : {};
                    for (var matName in mats) {
                        if (!Object.prototype.hasOwnProperty.call(mats, matName)) continue;
                        var kg = mats[matName];
                        var opts = '';
                        for (var n = 0; n < names.length; n++) {
                            opts +=
                                '<option value="' +
                                escapeAttr(names[n]) +
                                '"' +
                                (names[n] === matName ? ' selected' : '') +
                                '>' +
                                escapeHtml(names[n]) +
                                '</option>';
                        }
                        if (names.indexOf(matName) < 0) {
                            opts =
                                '<option value="' +
                                escapeAttr(matName) +
                                '" selected>' +
                                escapeHtml(matName) +
                                '</option>' +
                                opts;
                        }
                        matsHtml +=
                            '<tr><td><select class="rec-mat-name">' +
                            opts +
                            '</select></td><td><input type="number" step="any" class="rec-mat-kg" value="' +
                            Number(kg) +
                            '" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-rec-row" title="Удалить">✕</button></td></tr>';
                    }
                    block.innerHTML =
                        '<h4>Состав</h4><div class="cfg-recipe-name"><input type="text" class="rec-name" value="' +
                        escapeAttr(rec.name || '') +
                        '" placeholder="Наименование позиции" /></div><div class="cfg-table-wrap"><table class="cfg-table"><thead><tr><th>Материал</th><th>кг</th><th class="col-del"></th></tr></thead><tbody>' +
                        matsHtml +
                        '</tbody></table></div><button type="button" class="cfg-btn-sm cfg-add-rec-row">+ Строка</button> <button type="button" class="cfg-btn-sm cfg-del-recipe" title="Удалить состав">Удалить состав</button>';
                    cfgRecipesList.appendChild(block);
                }
            }

            function renderPrices(arr) {
                arr = arr || [];
                if (!cfgPricesBody) return;
                cfgPricesBody.innerHTML = '';
                for (var i = 0; i < arr.length; i++) {
                    var p = arr[i] || {};
                    var tr = document.createElement('tr');
                    tr.setAttribute('data-idx', String(i));
                    tr.innerHTML =
                        '<td><input type="text" class="price-name" value="' +
                        escapeAttr(p.name || '') +
                        '" /></td><td><input type="number" step="any" class="price-nd-nv" value="' +
                        (p.no_delivery_no_vat != null ? p.no_delivery_no_vat : '') +
                        '" /></td><td><input type="number" step="any" class="price-nd-v" value="' +
                        (p.no_delivery_vat_22 != null ? p.no_delivery_vat_22 : '') +
                        '" /></td><td><input type="number" step="any" class="price-pick-nv" value="' +
                        (p.pickup_no_vat != null ? p.pickup_no_vat : '') +
                        '" /></td><td><input type="number" step="any" class="price-pick-v" value="' +
                        (p.pickup_vat_22 != null ? p.pickup_vat_22 : '') +
                        '" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-price" title="Удалить">✕</button></td>';
                    cfgPricesBody.appendChild(tr);
                }
            }

            function getMaterialsFromUI() {
                var out = [];
                if (!cfgMaterialsBody) return out;
                var rows = cfgMaterialsBody.querySelectorAll('tr');
                for (var i = 0; i < rows.length; i++) {
                    var nameEl = rows[i].querySelector('.mat-name');
                    var name = nameEl ? String(nameEl.value || '').trim() : '';
                    var aliasesEl = rows[i].querySelector('.mat-aliases');
                    var aliasesStr = aliasesEl ? String(aliasesEl.value || '').trim() : '';
                    var aliases = [];
                    if (aliasesStr) {
                        var parts = aliasesStr.split('||');
                        for (var j = 0; j < parts.length; j++) {
                            var s = String(parts[j] || '').trim();
                            if (s) aliases.push(s);
                        }
                    }
                    out.push({ name: name, aliases: aliases });
                }
                return out;
            }

            function getRecipesFromUI() {
                var out = [];
                if (!cfgRecipesList) return out;
                var blocks = cfgRecipesList.querySelectorAll('.cfg-recipe-block');
                for (var i = 0; i < blocks.length; i++) {
                    var nmEl = blocks[i].querySelector('.rec-name');
                    var name = nmEl ? String(nmEl.value || '').trim() : '';
                    var materials = {};
                    var rows = blocks[i].querySelectorAll('tbody tr');
                    for (var r = 0; r < rows.length; r++) {
                        var matEl = rows[r].querySelector('.rec-mat-name');
                        var kgEl = rows[r].querySelector('.rec-mat-kg');
                        var matName = matEl ? String(matEl.value || '').trim() : '';
                        var kg = kgEl ? parseFloat(kgEl.value) : NaN;
                        if (matName && !isNaN(kg)) materials[matName] = kg;
                    }
                    out.push({ name: name, materials: materials });
                }
                return out;
            }

            function getPricesFromUI() {
                var out = [];
                if (!cfgPricesBody) return out;
                var rows = cfgPricesBody.querySelectorAll('tr');
                for (var i = 0; i < rows.length; i++) {
                    var nmEl = rows[i].querySelector('.price-name');
                    var name = nmEl ? String(nmEl.value || '').trim() : '';
                    function readNum(cls) {
                        var el = rows[i].querySelector(cls);
                        var v = el ? String(el.value || '').trim() : '';
                        return v === '' ? null : parseFloat(v);
                    }
                    var row = { name: name };
                    var v1 = readNum('.price-nd-nv');
                    var v2 = readNum('.price-nd-v');
                    var v3 = readNum('.price-pick-nv');
                    var v4 = readNum('.price-pick-v');
                    if (v1 != null && !isNaN(v1)) row.no_delivery_no_vat = v1;
                    if (v2 != null && !isNaN(v2)) row.no_delivery_vat_22 = v2;
                    if (v3 != null && !isNaN(v3)) row.pickup_no_vat = v3;
                    if (v4 != null && !isNaN(v4)) row.pickup_vat_22 = v4;
                    out.push(row);
                }
                return out;
            }

            function loadConfig(password) {
                if (!window.fetch) return;
                clearConfigErrors();
                setScopeTexts();
                return configFetch('/api/config?scope=' + encodeURIComponent(currentConfigScope), {}, password)
                    .then(function(res) {
                        if (!res.ok) {
                            return parseResponseError(res, 'Не удалось открыть конфигуратор.').then(function(error) {
                                throw new Error(error.message);
                            });
                        }
                        return res.json();
                    })
                    .then(function(data) {
                        if (sel) {
                            sel.innerHTML = '<option value="__base__">Базовая конфигурация</option>';
                            var profiles = data.profiles || [];
                            for (var i = 0; i < profiles.length; i++) {
                                var opt = document.createElement('option');
                                opt.value = profiles[i].name;
                                opt.textContent = profiles[i].name;
                                sel.appendChild(opt);
                            }
                            sel.value = data.active_profile || '__base__';
                        }
                        renderMaterials(data.materials || []);
                        currentExternalMaterialNames = data.external_materials || [];
                        var mats = data.materials || [];
                        var names = [];
                        for (var i2 = 0; i2 < mats.length; i2++) names.push(mats[i2].name);
                        for (var i3 = 0; i3 < currentExternalMaterialNames.length; i3++) names.push(currentExternalMaterialNames[i3]);
                        renderRecipes(data.recipes || [], names);
                        renderPrices(data.prices || []);
                        loadMainProfileSelects();
                        clearConfigStatus();
                        cfgPanel.style.display = 'flex';
                    })
                    .catch(function(e) {
                        closePanel();
                        alert('Ошибка доступа к конфигуратору: ' + (e && e.message ? e.message : String(e)));
                    });
            }

            function safeClosest(el, selector) {
                if (!el) return null;
                if (el.closest) return el.closest(selector);
                return null;
            }

            if (saveBtn) {
                saveBtn.addEventListener('click', function() {
                    clearConfigErrors();
                    reindexConfigUI();
                    var name = nameInput ? String(nameInput.value || '').trim() : '';
                    if (!name) {
                        if (nameInput) nameInput.classList.add('cfg-error-input');
                        showConfigStatus('Введите имя профиля, чтобы сохранить настройки.');
                        return;
                    }
                    if (!window.fetch) return;
                    var body = {
                        name: name,
                        scope: currentConfigScope,
                        recipes: getRecipesFromUI(),
                        prices: getPricesFromUI(),
                        materials: getMaterialsFromUI(),
                    };
                    configFetch('/api/config/profile', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify(body),
                    })
                        .then(function(res) {
                            if (res.ok) {
                                clearConfigErrors();
                                return null;
                            }
                            return parseResponseError(res, 'Не удалось сохранить профиль.').then(function(error) {
                                if (error.errors && error.errors.length) {
                                    applyValidationErrors(error.errors);
                                } else {
                                    showConfigStatus(error.message);
                                }
                                throw new Error(error.message);
                            });
                        })
                        .then(function() {
                            var p = loadConfig();
                            if (p && p.then && sel) {
                                p.then(function() {
                                    sel.value = name;
                                });
                            } else if (sel) {
                                sel.value = name;
                            }
                        })
                        .catch(function(e) {
                            if (e && e.message === 'Ошибка валидации конфигурации') return;
                            showConfigStatus(e && e.message ? e.message : String(e));
                        });
                });
            }

            if (applyBtn) {
                applyBtn.addEventListener('click', function() {
                    if (!window.fetch) return;
                    var name = sel ? sel.value : '__base__';
                    configFetch('/api/config/profile/select', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ name: name, scope: currentConfigScope }),
                    })
                        .then(function(res) {
                            if (res.ok) {
                                clearConfigErrors();
                                return null;
                            }
                            return parseResponseError(res, 'Не удалось выбрать профиль.').then(function(error) {
                                throw new Error(error.message);
                            });
                        })
                        .then(function() {
                            loadConfig();
                        })
                        .catch(function(e) {
                            showConfigStatus(e && e.message ? e.message : String(e));
                        });
                });
            }

            if (delBtn) {
                delBtn.addEventListener('click', function() {
                    if (!window.fetch) return;
                    var name = sel ? sel.value : '__base__';
                    if (name === '__base__') {
                        showConfigStatus('Базовую конфигурацию удалить нельзя.');
                        return;
                    }
                    if (!confirm('Удалить набор настроек "' + name + '"?')) return;
                    configFetch('/api/config/profile/' + encodeURIComponent(name) + '?scope=' + encodeURIComponent(currentConfigScope), { method: 'DELETE' })
                        .then(function(res) {
                            if (res.ok) {
                                clearConfigErrors();
                                return null;
                            }
                            return parseResponseError(res, 'Не удалось удалить профиль.').then(function(error) {
                                throw new Error(error.message);
                            });
                        })
                        .then(function() {
                            loadConfig();
                        })
                        .catch(function(e) {
                            showConfigStatus(e && e.message ? e.message : String(e));
                        });
                });
            }

            var cfgMaterialsAdd = document.getElementById('cfgMaterialsAdd');
            var cfgPricesAdd = document.getElementById('cfgPricesAdd');
            var cfgRecipesAdd = document.getElementById('cfgRecipesAdd');
            if (cfgMaterialsAdd && cfgMaterialsBody) {
                cfgMaterialsAdd.addEventListener('click', function() {
                    var tr = document.createElement('tr');
                    tr.innerHTML = '<td><input type="text" class="mat-name" /></td><td><input type="text" class="mat-aliases" placeholder="через ||" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-mat" title="Удалить">✕</button></td>';
                    cfgMaterialsBody.appendChild(tr);
                });
            }
            if (cfgPricesAdd && cfgPricesBody) {
                cfgPricesAdd.addEventListener('click', function() {
                    var tr = document.createElement('tr');
                    tr.innerHTML = '<td><input type="text" class="price-name" /></td><td><input type="number" step="any" class="price-nd-nv" /></td><td><input type="number" step="any" class="price-nd-v" /></td><td><input type="number" step="any" class="price-pick-nv" /></td><td><input type="number" step="any" class="price-pick-v" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-price" title="Удалить">✕</button></td>';
                    cfgPricesBody.appendChild(tr);
                });
            }
            if (cfgRecipesAdd && cfgRecipesList) {
                cfgRecipesAdd.addEventListener('click', function() {
                    var materialNames = getAvailableMaterialNames();
                    var opts = '';
                    if (materialNames.length) {
                        for (var j = 0; j < materialNames.length; j++) {
                            opts += '<option value="' + escapeAttr(materialNames[j]) + '">' + escapeHtml(materialNames[j]) + '</option>';
                        }
                    } else {
                        opts = '<option value="">— выберите материал —</option>';
                    }
                    var block = document.createElement('div');
                    block.className = 'cfg-recipe-block';
                    block.innerHTML = '<h4>Состав</h4><div class="cfg-recipe-name"><input type="text" class="rec-name" placeholder="Наименование позиции" /></div><div class="cfg-table-wrap"><table class="cfg-table"><thead><tr><th>Материал</th><th>кг</th><th class="col-del"></th></tr></thead><tbody><tr><td><select class="rec-mat-name">' + opts + '</select></td><td><input type="number" step="any" class="rec-mat-kg" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-rec-row" title="Удалить">✕</button></td></tr></tbody></table></div><button type="button" class="cfg-btn-sm cfg-add-rec-row">+ Строка</button> <button type="button" class="cfg-btn-sm cfg-del-recipe" title="Удалить состав">Удалить состав</button>';
                    cfgRecipesList.appendChild(block);
                });
            }

            if (cfgMaterialsBody) cfgMaterialsBody.addEventListener('click', function(e) {
                if (e.target && e.target.classList && e.target.classList.contains('cfg-del-mat')) {
                    var tr = safeClosest(e.target, 'tr');
                    if (tr && tr.parentNode) tr.parentNode.removeChild(tr);
                }
            });
            if (cfgPricesBody) cfgPricesBody.addEventListener('click', function(e) {
                if (e.target && e.target.classList && e.target.classList.contains('cfg-del-price')) {
                    var tr = safeClosest(e.target, 'tr');
                    if (tr && tr.parentNode) tr.parentNode.removeChild(tr);
                }
            });
            if (cfgRecipesList) cfgRecipesList.addEventListener('click', function(e) {
                if (!e.target || !e.target.classList) return;
                if (e.target.classList.contains('cfg-del-recipe')) {
                    var b = safeClosest(e.target, '.cfg-recipe-block');
                    if (b && b.parentNode) b.parentNode.removeChild(b);
                }
                if (e.target.classList.contains('cfg-del-rec-row')) {
                    var tr = safeClosest(e.target, 'tr');
                    if (tr && tr.parentNode) tr.parentNode.removeChild(tr);
                }
                if (e.target.classList.contains('cfg-add-rec-row')) {
                    var block = safeClosest(e.target, '.cfg-recipe-block');
                    var tbody = block ? block.querySelector('tbody') : null;
                    var materialNames2 = getAvailableMaterialNames();
                    var opts2 = '';
                    if (materialNames2.length) {
                        for (var j = 0; j < materialNames2.length; j++) {
                            opts2 += '<option value="' + escapeAttr(materialNames2[j]) + '">' + escapeHtml(materialNames2[j]) + '</option>';
                        }
                    } else {
                        opts2 = '<option value="">— выберите —</option>';
                    }
                    var row = document.createElement('tr');
                    row.innerHTML = '<td><select class="rec-mat-name">' + opts2 + '</select></td><td><input type="number" step="any" class="rec-mat-kg" /></td><td class="col-del"><button type="button" class="cfg-btn-sm cfg-del-rec-row" title="Удалить">✕</button></td>';
                    if (tbody) tbody.appendChild(row);
                }
            });

            function destroyPreviewCharts() {
                for (var pi = 0; pi < __previewChartInstances.length; pi++) {
                    try {
                        if (__previewChartInstances[pi] && typeof __previewChartInstances[pi].destroy === 'function') {
                            __previewChartInstances[pi].destroy();
                        }
                    } catch (ePrev) {}
                }
                __previewChartInstances.length = 0;
            }
            function shortChartLabel(s, maxLen) {
                s = String(s || '');
                if (s.length <= maxLen) return s;
                return s.slice(0, maxLen - 1) + '…';
            }
            function initBetonPreviewCharts(items) {
                if (typeof Chart === 'undefined' || !items || !items.length) return;
                destroyPreviewCharts();
                var elP = document.getElementById('betonPreviewPrice');
                var elV = document.getElementById('betonPreviewVol');
                if (!elP || !elV) return;
                var labels = items.map(function(it) { return shortChartLabel(it.name || '', 26); });
                var keys = ['no_delivery_no_vat', 'no_delivery_vat_22', 'pickup_no_vat', 'pickup_vat_22'];
                var keyLabels = ['Без доставки без НДС', 'С НДС 22%', 'Самовывоз без НДС', 'Самовывоз НДС 22%'];
                var palette = [
                    'rgba(37, 99, 235, 0.82)',
                    'rgba(8, 145, 178, 0.82)',
                    'rgba(217, 119, 6, 0.82)',
                    'rgba(185, 28, 28, 0.78)'
                ];
                var borderPalette = [
                    'rgb(37, 99, 235)',
                    'rgb(8, 145, 178)',
                    'rgb(217, 119, 6)',
                    'rgb(185, 28, 28)'
                ];
                var ds = [];
                for (var ki = 0; ki < keys.length; ki++) {
                    var arr = [];
                    for (var ri = 0; ri < items.length; ri++) {
                        var up = (items[ri].unit_prices || {})[keys[ki]];
                        arr.push(up != null && !isNaN(up) ? Number(up) : 0);
                    }
                    ds.push({
                        label: keyLabels[ki],
                        data: arr,
                        backgroundColor: palette[ki],
                        borderColor: borderPalette[ki],
                        borderWidth: 1,
                        borderRadius: 5
                    });
                }
                Chart.defaults.font.family = "'Segoe UI', system-ui, sans-serif";
                Chart.defaults.color = '#475569';
                var ch1 = new Chart(elP, {
                    type: 'bar',
                    data: { labels: labels, datasets: ds },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {
                            legend: { position: 'bottom', labels: { boxWidth: 10, font: { size: 10 }, padding: 10 } },
                            title: { display: false }
                        },
                        scales: {
                            x: { ticks: { maxRotation: 45, autoSkip: true } },
                            y: { beginAtZero: true, title: { display: true, text: '₽ / м³' } }
                        }
                    }
                });
                __previewChartInstances.push(ch1);
                var vols = items.map(function(it) { return Number(it.max_m3) || 0; });
                var ch2 = new Chart(elV, {
                    type: 'bar',
                    data: {
                        labels: labels,
                        datasets: [{
                            label: 'Макс. объём, м³',
                            data: vols,
                            backgroundColor: 'rgba(59, 130, 246, 0.75)',
                            borderColor: 'rgb(37, 99, 235)',
                            borderWidth: 1,
                            borderRadius: 6
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: false } },
                        scales: {
                            x: { ticks: { maxRotation: 45, autoSkip: true } },
                            y: { beginAtZero: true, title: { display: true, text: 'м³' } }
                        }
                    }
                });
                __previewChartInstances.push(ch2);
            }
            function initJbiPreviewCharts(items) {
                if (typeof Chart === 'undefined' || !items || !items.length) return;
                destroyPreviewCharts();
                var elP = document.getElementById('jbiPreviewPrice');
                var elV = document.getElementById('jbiPreviewVol');
                if (!elP || !elV) return;
                var labels = items.map(function(it) { return shortChartLabel(it.name || '', 26); });
                Chart.defaults.font.family = "'Segoe UI', system-ui, sans-serif";
                Chart.defaults.color = '#475569';
                var prices = items.map(function(it) { return Number(it.unit_price) || 0; });
                var ch1 = new Chart(elP, {
                    type: 'bar',
                    data: {
                        labels: labels,
                        datasets: [{
                            label: 'Цена за 1 шт',
                            data: prices,
                            backgroundColor: 'rgba(14, 116, 144, 0.78)',
                            borderColor: 'rgb(14, 116, 144)',
                            borderWidth: 1,
                            borderRadius: 6
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: false } },
                        scales: {
                            x: { ticks: { maxRotation: 45, autoSkip: true } },
                            y: { beginAtZero: true, title: { display: true, text: '₽' } }
                        }
                    }
                });
                __previewChartInstances.push(ch1);
                var units = items.map(function(it) { return Number(it.max_units) || 0; });
                var ch2 = new Chart(elV, {
                    type: 'bar',
                    data: {
                        labels: labels,
                        datasets: [{
                            label: 'Макс. шт',
                            data: units,
                            backgroundColor: 'rgba(37, 99, 235, 0.75)',
                            borderColor: 'rgb(37, 99, 235)',
                            borderWidth: 1,
                            borderRadius: 6
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: false } },
                        scales: {
                            x: { ticks: { maxRotation: 45, autoSkip: true } },
                            y: { beginAtZero: true, title: { display: true, text: 'шт' } }
                        }
                    }
                });
                __previewChartInstances.push(ch2);
            }

            // основная форма расчета: отдельные действия "Посчитать" и "Скачать"
            if (calcForm && window.fetch && resultBox && calcOnlyBtn && downloadBtn) {
                function runBetonCalculation(shouldDownload) {
                    var fileInput = document.getElementById('file');
                    if (!fileInput || !fileInput.files || !fileInput.files[0]) {
                        alert('Выберите файл .xlsx');
                        return;
                    }
                    calcOnlyBtn.disabled = true;
                    downloadBtn.disabled = true;
                    calcOnlyBtn.textContent = shouldDownload ? 'Посчитать' : 'Считаем...';
                    downloadBtn.textContent = shouldDownload ? 'Скачиваем...' : 'Скачать';
                    resultBox.classList.remove('result-ok');
                    resultBox.classList.remove('result-empty');
                    resultBox.classList.remove('has-result');
                    resultBox.innerHTML = '<div class="result-title">Результат расчета</div><div class="result-meta">Выполняется расчет, подождите...</div>';

                    var fd = new FormData(calcForm);
                    fd.set('mode', shouldDownload ? 'excel' : 'summary');
                    fd.set('scope', 'beton');
                    fd.set('profile_name', mainProfileSelect ? mainProfileSelect.value : '__base__');

                    function renderBetonSummary(result) {
                            var items = result.items || [];

                            var html = '<div class="result-title">Результат расчета</div>';

                            function fmtVolume(v) {
                                return formatNumber(v, 3);
                            }
                            function fmtMoney(v) {
                                return v != null && !isNaN(v) ? formatNumber(v, 2) + ' ₽' : '—';
                            }
                            function formatLimiterText(limiters) {
                                if (!limiters || !limiters.length) return '';
                                var names = [];
                                for (var li = 0; li < limiters.length; li++) {
                                    if (limiters[li] && limiters[li].material) names.push(limiters[li].material);
                                }
                                if (!names.length) return '';
                                return (names.length > 1 ? 'Ограничивают: ' : 'Ограничивает: ') + names.join(', ');
                            }
                            function getMaxIndex(sourceItems, key) {
                                var bestIdx = -1;
                                var bestVal = -1;
                                for (var ii = 0; ii < sourceItems.length; ii++) {
                                    var amounts = sourceItems[ii].amounts || {};
                                    var value = amounts[key];
                                    if (value != null && !isNaN(value) && value > bestVal) {
                                        bestVal = value;
                                        bestIdx = ii;
                                    }
                                }
                                return bestIdx;
                            }
                            function buildTable(title, colA, colB, maxKey, labelA, labelB) {
                                var maxIdx = getMaxIndex(items, maxKey);
                                var out = '<div class="result-table-card">';
                                out += '<div class="result-section-title">' + escapeHtml(title) + '</div>';
                                out += '<div class="result-section-subtitle">Объем и цены по каждому виду бетона</div>';
                                out += '<div class="result-table-wrap"><table class="result-table">';
                                out += '<thead><tr><th>Наименование бетона</th><th class="result-num">Объем, м³</th><th class="result-num">' + escapeHtml(labelA) + '</th><th class="result-num">' + escapeHtml(labelB) + '</th></tr></thead><tbody>';
                                for (var iii = 0; iii < items.length; iii++) {
                                    var item = items[iii];
                                    var amounts = item.amounts || {};
                                    var limiterText = Number(item.max_m3 || 0) === 0 ? formatLimiterText(item.limiters || []) : '';
                                    var isMax = iii === maxIdx;
                                    out += '<tr' + (isMax ? ' class="row-max"' : '') + '>';
                                    out += '<td><span class="result-name">' + escapeHtml(item.name || '') + '</span>' + (isMax ? '<span class="result-badge">Макс. цена</span>' : '') + (limiterText ? '<div class="result-limiter">' + escapeHtml(limiterText) + '</div>' : '') + '</td>';
                                    out += '<td class="result-num">' + fmtVolume(item.max_m3) + '</td>';
                                    out += '<td class="result-num">' + fmtMoney(amounts[colA]) + '</td>';
                                    out += '<td class="result-num">' + fmtMoney(amounts[colB]) + '</td>';
                                    out += '</tr>';
                                }
                                out += '</tbody></table></div></div>';
                                return out;
                            }

                            if (items.length) {
                                html += '<div class="result-tables">';
                                html += buildTable(
                                    'Для организации А',
                                    'no_delivery_no_vat',
                                    'no_delivery_vat_22',
                                    'no_delivery_vat_22',
                                    'Без доставки без НДС',
                                    'Без доставки с НДС 22%'
                                );
                                html += buildTable(
                                    'Для иных организаций',
                                    'pickup_no_vat',
                                    'pickup_vat_22',
                                    'pickup_vat_22',
                                    'Самовывоз без НДС',
                                    'Самовывоз с НДС 22%'
                                );
                                html += '</div>';
                                html += '<div class="chart-preview">';
                                html += '<div class="chart-preview-head">';
                                html += '<div class="chart-preview-title">Графики (пример)</div>';
                                html += '<div class="chart-preview-hint">Макет для будущей аналитики. Сейчас подставляются данные этого расчёта: сравнение цен за 1 м³ по сценариям и объёмы по видам бетона.</div>';
                                html += '</div>';
                                html += '<div class="chart-preview-grid">';
                                html += '<div class="chart-preview-card"><div class="chart-preview-cap">Цена за 1 м³ — четыре сценария</div><div class="chart-preview-canvas"><canvas id="betonPreviewPrice"></canvas></div></div>';
                                html += '<div class="chart-preview-card"><div class="chart-preview-cap">Допустимый объём по видам</div><div class="chart-preview-canvas"><canvas id="betonPreviewVol"></canvas></div></div>';
                                html += '</div></div>';
                            } else {
                                html +=
                                    '<ul class="result-list"><li>Данные по бетонам отсутствуют. Проверьте исходный файл.</li></ul>';
                            }

                            resultBox.classList.add('result-ok');
                            resultBox.classList.add('has-result');
                            destroyPreviewCharts();
                            resultBox.innerHTML = html;
                            if (items.length) {
                                requestAnimationFrame(function() { initBetonPreviewCharts(items); });
                            }

                    }

                    function doDownload(jobId) {
                        return fetch('/upload/file/' + encodeURIComponent(jobId)).then(function(res) {
                            if (!res.ok) throw new Error('Ошибка загрузки файла');
                            return res.blob().then(function(blob) {
                                var url = window.URL.createObjectURL(blob);
                                var a = document.createElement('a');
                                a.href = url;
                                a.download = 'raschet_po_ostatkam.xlsx';
                                document.body.appendChild(a);
                                a.click();
                                setTimeout(function() {
                                    document.body.removeChild(a);
                                    window.URL.revokeObjectURL(url);
                                }, 0);
                            });
                        });
                    }

                    function pollResult(jobId) {
                        return new Promise(function(resolve, reject) {
                            var iv = setInterval(function() {
                                fetch('/upload/result/' + encodeURIComponent(jobId))
                                    .then(function(r) { return r.json(); })
                                    .then(function(data) {
                                        if (data.status === 'ready') {
                                            clearInterval(iv);
                                            renderBetonSummary(data.summary);
                                            if (shouldDownload && data.has_excel) {
                                                doDownload(jobId).then(resolve).catch(reject);
                                            } else {
                                                resolve();
                                            }
                                        } else if (data.status === 'failed') {
                                            clearInterval(iv);
                                            reject(new Error(data.error || 'Ошибка обработки'));
                                        }
                                    })
                                    .catch(function(err) {
                                        clearInterval(iv);
                                        reject(err);
                                    });
                            }, 500);
                        });
                    }

                    fetch('/upload', { method: 'POST', body: fd })
                        .then(function(res) {
                            if (!res.ok) {
                                return res.text().then(function(t) { throw new Error(t || 'Ошибка сервера'); });
                            }
                            var ct = (res.headers.get('content-type') || '');
                            if (ct.indexOf('json') >= 0) {
                                return res.json();
                            }
                            return res.blob().then(function(blob) {
                                var url = window.URL.createObjectURL(blob);
                                var a = document.createElement('a');
                                a.href = url;
                                a.download = 'raschet_po_ostatkam.xlsx';
                                document.body.appendChild(a);
                                a.click();
                                setTimeout(function() {
                                    document.body.removeChild(a);
                                    window.URL.revokeObjectURL(url);
                                }, 0);
                                resultBox.classList.add('result-ok');
                                resultBox.classList.add('has-result');
                                resultBox.innerHTML = '<div class="result-title">Результат расчета</div><div class="result-meta">Файл скачан.</div>';
                                return null;
                            });
                        })
                        .then(function(result) {
                            if (!result) return;
                            if (result.job_id) {
                                return pollResult(result.job_id);
                            }
                            renderBetonSummary(result);
                        })
                        .catch(function(err) {
                            resultBox.classList.remove('result-ok');
                            resultBox.classList.remove('has-result');
                            resultBox.innerHTML =
                                '<div class="result-title">Ошибка расчета</div><div class="result-meta">' +
                                escapeHtml(err && err.message ? err.message : String(err)) +
                                '</div>';
                        })
                        .finally(function() {
                            calcOnlyBtn.disabled = false;
                            downloadBtn.disabled = false;
                            calcOnlyBtn.textContent = 'Посчитать';
                            downloadBtn.textContent = 'Скачать';
                        });
                }

                calcOnlyBtn.addEventListener('click', function() {
                    runBetonCalculation(false);
                });
                downloadBtn.addEventListener('click', function() {
                    runBetonCalculation(true);
                });
            }

            if (jbiForm && resultBox && jbiCalcOnlyBtn && jbiDownloadBtn) {
                function runJbiCalculation(shouldDownload) {
                    var jbiFileInput = document.getElementById('jbiFile');
                    if (!jbiFileInput || !jbiFileInput.files || !jbiFileInput.files[0]) {
                        alert('Выберите файл .xlsx');
                        return;
                    }
                    jbiCalcOnlyBtn.disabled = true;
                    jbiDownloadBtn.disabled = true;
                    jbiCalcOnlyBtn.textContent = shouldDownload ? 'Посчитать' : 'Считаем...';
                    jbiDownloadBtn.textContent = shouldDownload ? 'Скачиваем...' : 'Скачать';
                    resultBox.classList.remove('result-ok');
                    resultBox.classList.remove('result-empty');
                    resultBox.classList.remove('has-result');
                    resultBox.innerHTML = '<div class="result-title">Результат расчета</div><div class="result-meta">Выполняется расчет ЖБИ, подождите...</div>';

                    var fdJbi = new FormData(jbiForm);
                    fdJbi.set('mode', shouldDownload ? 'excel' : 'summary');
                    fdJbi.set('scope', 'jbi');
                    fdJbi.set('profile_name', jbiProfileSelect ? jbiProfileSelect.value : '__base__');

                    function renderJbiSummary(summary) {
                            var items = summary.items || [];
                            var html = '<div class="result-title">Результат расчета</div>';
                            function fmtMoney(v) {
                                return v != null && !isNaN(v) ? formatNumber(v, 2) + ' ₽' : '—';
                            }
                            function formatLimiterText(limiters) {
                                if (!limiters || !limiters.length) return '';
                                var names = [];
                                for (var li = 0; li < limiters.length; li++) {
                                    if (limiters[li] && limiters[li].material) names.push(limiters[li].material);
                                }
                                if (!names.length) return '';
                                return (names.length > 1 ? 'Ограничивают: ' : 'Ограничивает: ') + names.join(', ');
                            }
                            if (items.length) {
                                html += '<div class="result-table-card">';
                                html += '<div class="result-section-title">Расчет ЖБИ</div>';
                                html += '<div class="result-section-subtitle">Максимум изделий и итоговая стоимость</div>';
                                html += '<div class="result-table-wrap"><table class="result-table">';
                                html += '<thead><tr><th>Наименование изделия</th><th class="result-num">Максимум, шт</th><th class="result-num">Цена за 1 шт</th><th class="result-num">Общая цена</th></tr></thead><tbody>';
                                for (var i = 0; i < items.length; i++) {
                                    var limiterText = Number(items[i].max_units || 0) === 0 ? formatLimiterText(items[i].limiters || []) : '';
                                    html += '<tr><td><span class="result-name">' + escapeHtml(items[i].name || '') + '</span>' + (limiterText ? '<div class="result-limiter">' + escapeHtml(limiterText) + '</div>' : '') + '</td><td class="result-num">' + String(items[i].max_units != null ? items[i].max_units : 0) + '</td><td class="result-num">' + fmtMoney(items[i].unit_price) + '</td><td class="result-num">' + fmtMoney(items[i].total_price) + '</td></tr>';
                                }
                                html += '</tbody></table></div></div>';
                                html += '<div class="chart-preview">';
                                html += '<div class="chart-preview-head">';
                                html += '<div class="chart-preview-title">Графики (пример)</div>';
                                html += '<div class="chart-preview-hint">Макет для будущей аналитики. Сейчас — цена за 1 шт и максимум штук по изделиям из этого расчёта.</div>';
                                html += '</div>';
                                html += '<div class="chart-preview-grid">';
                                html += '<div class="chart-preview-card"><div class="chart-preview-cap">Цена за 1 шт</div><div class="chart-preview-canvas"><canvas id="jbiPreviewPrice"></canvas></div></div>';
                                html += '<div class="chart-preview-card"><div class="chart-preview-cap">Максимум изделий, шт</div><div class="chart-preview-canvas"><canvas id="jbiPreviewVol"></canvas></div></div>';
                                html += '</div></div>';
                            } else {
                                html += '<div class="result-meta">Данные по ЖБИ отсутствуют.</div>';
                            }
                            resultBox.classList.add('result-ok');
                            resultBox.classList.add('has-result');
                            destroyPreviewCharts();
                            resultBox.innerHTML = html;
                            if (items.length) {
                                requestAnimationFrame(function() { initJbiPreviewCharts(items); });
                            }
                    }
                    function doDownloadJbi(jobId) {
                        return fetch('/upload/file/' + encodeURIComponent(jobId)).then(function(res) {
                            if (!res.ok) throw new Error('Ошибка загрузки файла');
                            return res.blob().then(function(blob) {
                                var url = window.URL.createObjectURL(blob);
                                var a = document.createElement('a');
                                a.href = url;
                                a.download = 'raschet_zhb.xlsx';
                                document.body.appendChild(a);
                                a.click();
                                setTimeout(function() {
                                    document.body.removeChild(a);
                                    window.URL.revokeObjectURL(url);
                                }, 0);
                            });
                        });
                    }

                    function pollJbiResult(jobId) {
                        return new Promise(function(resolve, reject) {
                            var iv = setInterval(function() {
                                fetch('/upload/result/' + encodeURIComponent(jobId))
                                    .then(function(r) { return r.json(); })
                                    .then(function(data) {
                                        if (data.status === 'ready') {
                                            clearInterval(iv);
                                            renderJbiSummary(data.summary);
                                            if (shouldDownload && data.has_excel) {
                                                doDownloadJbi(jobId).then(resolve).catch(reject);
                                            } else {
                                                resolve();
                                            }
                                        } else if (data.status === 'failed') {
                                            clearInterval(iv);
                                            reject(new Error(data.error || 'Ошибка обработки'));
                                        }
                                    })
                                    .catch(function(err) {
                                        clearInterval(iv);
                                        reject(err);
                                    });
                            }, 500);
                        });
                    }
                    fetch('/upload', { method: 'POST', body: fdJbi })
                        .then(function(res) {
                            if (!res.ok) {
                                return res.text().then(function(t) { throw new Error(t || 'Ошибка расчета ЖБИ'); });
                            }
                            var ct = (res.headers.get('content-type') || '');
                            if (ct.indexOf('json') >= 0) {
                                return res.json();
                            }
                            return res.blob().then(function(blob) {
                                var url = window.URL.createObjectURL(blob);
                                var a = document.createElement('a');
                                a.href = url;
                                a.download = 'raschet_zhb.xlsx';
                                document.body.appendChild(a);
                                a.click();
                                setTimeout(function() {
                                    document.body.removeChild(a);
                                    window.URL.revokeObjectURL(url);
                                }, 0);
                                resultBox.classList.add('result-ok');
                                resultBox.classList.add('has-result');
                                resultBox.innerHTML = '<div class="result-title">Результат расчета</div><div class="result-meta">Файл скачан.</div>';
                                return null;
                            });
                        })
                        .then(function(result) {
                            if (!result) return;
                            if (result.job_id) {
                                return pollJbiResult(result.job_id);
                            }
                            renderJbiSummary(result);
                        })
                        .catch(function(err) {
                            resultBox.classList.remove('result-ok');
                            resultBox.classList.remove('has-result');
                            resultBox.innerHTML =
                                '<div class="result-title">Ошибка расчета</div><div class="result-meta">' +
                                escapeHtml(err && err.message ? err.message : String(err)) +
                                '</div>';
                        })
                        .finally(function() {
                            jbiCalcOnlyBtn.disabled = false;
                            jbiDownloadBtn.disabled = false;
                            jbiCalcOnlyBtn.textContent = 'Посчитать';
                            jbiDownloadBtn.textContent = 'Скачать';
                        });
                }

                jbiCalcOnlyBtn.addEventListener('click', function() {
                    runJbiCalculation(false);
                });
                jbiDownloadBtn.addEventListener('click', function() {
                    runJbiCalculation(true);
                });
            }
        });
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)


def _upload_sync(
    content: bytes,
    filename: str,
    mode: str,
    scope: str,
    selected_profile: Optional[str],
):
    """Синхронная обработка (fallback при недоступности очереди)."""
    direction = get_direction(scope)
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / (filename or "остатки.xlsx")
        input_path.write_bytes(content)
        materials = _load_materials(scope=scope, profile_name=selected_profile)
        concrete_materials = (
            _load_materials(scope=direction.concrete_source, profile_name=None)
            if direction.concrete_source
            else []
        )
        balances = extract_balances(str(input_path), materials)
        concrete_balances: dict[str, float] = {}
        if direction.concrete_source:
            try:
                concrete_balances = extract_balances(str(input_path), concrete_materials)
            except Exception:
                pass
        if direction.calc_type == "units":
            combined = {**balances, **concrete_balances}
            summary = _build_jbi_summary(combined, profile_name=selected_profile)
            if mode == "excel":
                excel_bytes = _build_jbi_excel(combined, profile_name=selected_profile)
                return {"summary": summary, "has_excel": True, "excel_bytes": excel_bytes}
            return {"summary": summary, "has_excel": False}
        summary = _build_summary(balances, scope=scope, profile_name=selected_profile)
        if mode == "excel":
            excel_bytes = _build_excel(balances, scope=scope, profile_name=selected_profile)
            return {"summary": summary, "has_excel": True, "excel_bytes": excel_bytes}
        return {"summary": summary, "has_excel": False}


@app.post("/upload")
async def upload(
    request: Request,
    file: UploadFile = File(...),
    website: str = Form(""),
    mode: str = Form("excel"),
    scope: str = Form("beton"),
    profile_name: str = Form("__base__"),
):
    if website:
        raise HTTPException(status_code=400, detail="Spam detected")

    scope = validate_scope(scope)
    selected_profile = None if profile_name == "__base__" else profile_name

    ip = _client_ip(request)
    if mode == "excel" and _is_rate_limited(ip):
        raise HTTPException(status_code=429, detail="Слишком много запросов, попробуйте позже.")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel .xlsx.")

    if not get_direction(scope).supports_excel and mode == "excel":
        raise HTTPException(status_code=400, detail="Excel для данного направления пока не реализован")

    content = await file.read()
    filename = file.filename or "остатки.xlsx"

    import os
    use_queue = not os.environ.get("TESTING")
    if use_queue:
        try:
            from app.tasks import process_excel_task
            task = process_excel_task.delay(
                content, filename, mode, scope, profile_name
            )
            return JSONResponse(content={"job_id": task.id})
        except Exception as e:
            from celery.exceptions import OperationalError
            if not isinstance(e, (ImportError, OperationalError)):
                raise
    result = _upload_sync(content, filename, mode, scope, selected_profile)
    if result.get("has_excel") and result.get("excel_bytes"):
        return StreamingResponse(
            io.BytesIO(result["excel_bytes"]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="raschet_po_ostatkam.xlsx"'},
        )
    return JSONResponse(content=result["summary"])


@app.get("/upload/result/{job_id}")
async def upload_result(job_id: str):
    """Получить результат фоновой обработки."""
    try:
        from celery.result import AsyncResult
        from app.celery_app import app as celery_app
        ar = AsyncResult(job_id, app=celery_app)
        if ar.state == "PENDING":
            return JSONResponse(content={"status": "pending"})
        if ar.state == "SUCCESS":
            result = ar.result
            if isinstance(result, dict) and "summary" in result:
                return JSONResponse(content={
                    "status": "ready",
                    "summary": result["summary"],
                    "has_excel": result.get("has_excel", False),
                })
            return JSONResponse(content={"status": "ready", "summary": result})
        if ar.state == "FAILURE":
            return JSONResponse(
                status_code=500,
                content={"status": "failed", "error": str(ar.result) if ar.result else "Ошибка обработки"},
            )
    except (ImportError, Exception) as e:
        return JSONResponse(status_code=503, content={"status": "error", "error": str(e)})
    return JSONResponse(content={"status": "pending"})


@app.get("/upload/file/{job_id}")
async def upload_file(job_id: str):
    """Скачать Excel-файл по job_id."""
    from app.tasks import JOBS_DIR
    path = JOBS_DIR / f"{job_id}.xlsx"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Файл не найден или устарел")
    return StreamingResponse(
        io.BytesIO(path.read_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="raschet_po_ostatkam.xlsx"'},
    )


@app.get("/api/directions")
async def api_get_directions() -> List[Dict[str, Any]]:
    """Список направлений расчёта для динамического UI."""
    return [
        {
            "id": d.id,
            "display_name": d.display_name,
            "calc_type": d.calc_type,
            "concrete_source": d.concrete_source,
            "supports_excel": d.supports_excel,
        }
        for d in get_all_directions()
    ]


@app.get("/api/config/options")
async def api_get_config_options(scope: str = "beton") -> Dict[str, Any]:
    scope = validate_scope(scope)
    profiles = _load_profiles(scope)
    active_name: Optional[str] = profiles.get("active")
    return {
        "profiles": [{"name": p.get("name", "")} for p in profiles.get("profiles", [])],
        "active_profile": active_name or "__base__",
        "scope": scope,
    }


@app.get("/api/config")
async def api_get_config(request: Request, scope: str = "beton") -> Dict[str, Any]:
    _require_config_password(request)
    scope = validate_scope(scope)
    profiles = _load_profiles(scope)
    active_name: Optional[str] = profiles.get("active")
    active_profile = _get_profile(profiles, active_name)

    base = _default_config_payload(scope)
    materials = active_profile.get("materials") if active_profile and "materials" in active_profile else base["materials"]
    recipes = active_profile.get("recipes") if active_profile and "recipes" in active_profile else base["recipes"]
    prices = active_profile.get("prices") if active_profile and "prices" in active_profile else base["prices"]

    return {
        "materials": materials,
        "recipes": recipes,
        "prices": prices,
        "profiles": [{"name": p.get("name", "")} for p in profiles.get("profiles", [])],
        "active_profile": active_name or "__base__",
        "scope": scope,
        "external_materials": (
            [r.name for r in _load_recipes(scope=src)]
            if (src := get_direction(scope).concrete_source)
            else []
        ),
    }


@app.post("/api/config/profile")
async def api_save_profile(
    request: Request, payload: Dict[str, Any] = Body(...)
) -> Dict[str, str]:
    _require_config_password(request)
    scope = validate_scope(payload.get("scope"))
    validated = _validate_and_prepare_profile(payload, scope=scope)
    name = validated["name"]

    profiles = _load_profiles(scope)
    prof_list = profiles.get("profiles", [])

    new_profile = {
        "name": name,
        "materials": validated["materials"],
        "recipes": validated["recipes"],
        "prices": validated["prices"],
    }

    replaced = False
    for idx, p in enumerate(prof_list):
        if p.get("name") == name:
            prof_list[idx] = new_profile
            replaced = True
            break
    if not replaced:
        prof_list.append(new_profile)

    profiles["profiles"] = prof_list
    profiles["active"] = name
    _save_profiles(profiles, scope)
    return {"status": "ok"}


@app.post("/api/config/profile/select")
async def api_select_profile(
    request: Request, payload: Dict[str, Any] = Body(...)
) -> Dict[str, str]:
    _require_config_password(request)
    scope = validate_scope(payload.get("scope"))
    name = (payload.get("name") or "").strip()

    profiles = _load_profiles(scope)
    if name == "__base__" or not name:
        profiles["active"] = None
        _save_profiles(profiles, scope)
        return {"status": "ok"}

    if not _get_profile(profiles, name):
        raise HTTPException(status_code=404, detail="Профиль не найден")

    profiles["active"] = name
    _save_profiles(profiles, scope)
    return {"status": "ok"}


@app.delete("/api/config/profile/{name}")
async def api_delete_profile(
    name: str, request: Request, scope: str = "beton"
) -> Dict[str, str]:
    _require_config_password(request)
    scope = validate_scope(scope)
    if name == "__base__":
        raise HTTPException(status_code=400, detail="Базовый профиль удалить нельзя")
    profiles = _load_profiles(scope)
    prof_list = profiles.get("profiles", [])
    prof_list = [p for p in prof_list if p.get("name") != name]
    profiles["profiles"] = prof_list
    if profiles.get("active") == name:
        profiles["active"] = None
    _save_profiles(profiles, scope)
    return {"status": "ok"}

